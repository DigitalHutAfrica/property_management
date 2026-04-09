# -*- coding: utf-8 -*-
# Copyright 2020-Today TechKhedut.
# Part of TechKhedut. See LICENSE file for full copyright and licensing details.
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import base64
from io import BytesIO
from odoo import fields, models, api, _
from odoo.exceptions import ValidationError, UserError


class PropertyDetails(models.Model):
    _name = 'property.details'
    _description = 'Property Details and for registration new Property'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    # Parent Property Field
    is_parent_property = fields.Boolean(string='Main Property', tracking=True)
    parent_property_id = fields.Many2one('parent.property', string='Property ',
                                         tracking=True)
    parent_amenities_ids = fields.Many2many(
        related='parent_property_id.amenities_ids', string="Parent Amenities",
        tracking=True)
    parent_specification_ids = fields.Many2many(
        related='parent_property_id.property_specification_ids',
        string="Parent Specification", tracking=True)

    # Parent Amenities
    parent_airport = fields.Char(
        related='parent_property_id.airport', string='Parent Airport ',
        tracking=True)
    parent_national_highway = fields.Char(
        related='parent_property_id.national_highway',
        string='Parent National Highway', tracking=True)
    parent_metro_station = fields.Char(
        related='parent_property_id.metro_station',
        string='Parent Metro Station ', tracking=True)
    parent_metro_city = fields.Char(
        related='parent_property_id.metro_city', string='Parent Metro City ',
        tracking=True)
    parent_school = fields.Char(
        related="parent_property_id.school", string="Parent School ",
        tracking=True)
    parent_hospital = fields.Char(
        related="parent_property_id.hospital", string="Hospital ",
        tracking=True)
    parent_shopping_mall = fields.Char(
        related="parent_property_id.shopping_mall", string="Parent Mall ",
        tracking=True)
    parent_park = fields.Char(
        related="parent_property_id.park", string="Parent Park ",
        tracking=True)
    # Address
    parent_zip = fields.Char(
        related='parent_property_id.zip', string='Parent Pin Code ',
        tracking=True)
    parent_street = fields.Char(
        related='parent_property_id.street', string='Parent Street1 ',
        tracking=True)
    parent_street2 = fields.Char(
        related='parent_property_id.street2', string='Parent Street2 ',
        tracking=True)
    parent_city_id = fields.Many2one(
        related='parent_property_id.city_id', string='Parent City ',
        tracking=True)
    parent_country_id = fields.Many2one(
        related='parent_property_id.country_id', string='Parent Country ',
        tracking=True)
    parent_state_id = fields.Many2one(
        related='parent_property_id.state_id', string='Parent State ',
        tracking=True)
    parent_website = fields.Char(
        related='parent_property_id.website', string='Parent Website ',
        tracking=True)

    # Common Field

    property_seq = fields.Char(string='Sequence', required=True, readonly=True,
                               copy=False,
                               default=lambda self: 'New')
    name = fields.Char(string='Name', required=True, tracking=True)
    image = fields.Binary(string='Image')
    type = fields.Selection([('land', 'Land'),
                             ('residential', 'Residential'),
                             ('commercial', 'Commercial'),
                             ('industrial', 'Industrial')
                             ], string='Property Type', tracking=True,
                            required=True)
    stage = fields.Selection([('draft', 'Draft'),
                              ('available', 'Available'),
                              ('on_lease', 'On Lease'),
                              ('sold', 'Sold'),
                              ('occupied_by_staff', 'Occupied By Staff'),
                              ('booked', 'Booked')],
                             string='Stage', default='draft', required=True,
                             tracking=True, readonly=True)
    company_id = fields.Many2one(
        'res.company', string='Default Company',
        default=lambda self: self.env.company)
    company_currency_id = fields.Many2one(
        'res.currency', related='company_id.currency_id', string='Company Currency')
    longitude = fields.Char(string='Longitude')
    latitude = fields.Char(string='Latitude')
    sale_lease = fields.Selection([('for_sale', 'For Sale'),
                                   ('for_tenancy', 'For Rent')],
                                  string='Property For', default='for_tenancy',
                                  required=True, tracking=True)
    sale_price = fields.Monetary(string='Sale Price', tracking=True)
    tenancy_price = fields.Monetary(string='Rent', tracking=True)
    rent_unit = fields.Selection(
        [('Day', "Day"), ('Month', "Month"), ('Year', "Year"),
        ('square_meter', "Square Meter")], default='Month',
        string="Rent Unit", tracking=True)
    is_extra_service = fields.Boolean(string="Any Extra Services")
    token_amount = fields.Monetary(string='Booking Price', tracking=True)
    website = fields.Char(string='Website', tracking=True)
    sold_invoice_id = fields.Many2one(
        'account.move', string='Sold Invoice', readonly=True, tracking=True)
    sold_invoice_state = fields.Boolean(string='Sold Invoice State',
                                        tracking=True)
    construct_year = fields.Char(string="Construct Year", size=4,
                                 tracking=True)
    buying_year = fields.Char(string="Buying Year", size=4, tracking=True)
    property_licence_no = fields.Char(string='Licence No.', tracking=True)
    address = fields.Char()

    # Address
    zip = fields.Char(string='Pin Code', size=6, tracking=True)
    street = fields.Char(string='Street1', tracking=True)
    street2 = fields.Char(string='Street2', tracking=True)
    city = fields.Char(string='City Name', tracking=True)
    city_id = fields.Many2one('property.res.city', tracking=True,
                              string='City')
    country_id = fields.Many2one('res.country', 'Country', tracking=True)
    state_id = fields.Many2one(
        "res.country.state", tracking=True, string='State', readonly=False,
        store=True,
        domain="[('country_id', '=?', country_id)]")

    # Nearby Connectivity
    airport = fields.Char(string='Airport')
    national_highway = fields.Char(string='National Highway ')
    metro_station = fields.Char(string='Metro Station')
    metro_city = fields.Char(string='Metro City')
    school = fields.Char(string="School")
    hospital = fields.Char(string="Hospital")
    shopping_mall = fields.Char(string="Mall")
    park = fields.Char(string="Park")

    # Related Field
    landlord_id = fields.Many2one('res.partner', string='Landlord',
                                  domain=lambda self: [
                                      ('is_landlord', '=', True)],
                                  tracking=True)

    tenancy_ids = fields.One2many(
        'tenancy.details', 'property_id', string='History', tracking=True)
    broker_ids = fields.One2many('tenancy.details', 'property_id',
                                 string='Broker',
                                 domain=[('is_any_broker', '=', True)],
                                 tracking=True)
    amenities_ids = fields.Many2many('property.amenities', string='Amenities ',
                                     tracking=True)
    property_specification_ids = fields.Many2many(
        'property.specification', string='Specification ', tracking=True)
    property_vendor_ids = fields.One2many(
        'property.vendor', 'property_id', string='Vendor Details',
        tracking=True)
    certificate_ids = fields.One2many(
        'property.certificate', 'property_id', string='Certificate',
        tracking=True)
    maintenance_ids = fields.One2many(
        'maintenance.request', 'property_id', string='Maintenance History',
        tracking=True)
    floreplan_ids = fields.One2many(
        'floor.plan', 'property_id', string='Floor Plan', tracking=True)
    property_images_ids = fields.One2many(
        'property.images', 'property_id', string='Images', tracking=True)
    tag_ids = fields.Many2many('property.tag', string='Tags', tracking=True)
    extra_service_ids = fields.One2many(
        'extra.service.line', 'property_id', string="Extra Services",
        tracking=True)
    extra_service_cost = fields.Monetary(
        string="Total", compute="_compute_extra_service_cost", tracking=True)
    nearby_connectivity_ids = fields.Many2many(
        'property.connectivity', string="Nearby Connectivity ", tracking=True)
    sold_booking_id = fields.Many2one('property.vendor', string="Booking",
                                      tracking=True)
    tenancy_inquiry_ids = fields.One2many(
        'tenancy.inquiry', 'property_id', string="Tenancy Inquiry",
        tracking=True)
    sale_inquiry_ids = fields.One2many(
        'sale.inquiry', 'property_id', string="Sale Inquiry", tracking=True)
    contract_history_ids = fields.One2many(
        "tenancy.details", "property_id", string="Contract History",
        tracking=True)

    # Residential
    residence_type = fields.Selection([('apartment', 'Apartment'),
                                       ('bungalow', 'Bungalow'),
                                       ('vila', 'Vila'),
                                       ('raw_house', 'Raw House'),
                                       ('duplex', 'Duplex House'),
                                       ('single_studio', 'Single Studio')],
                                      string='Type of Residence',
                                      tracking=True)
    total_floor = fields.Integer(string='Total Floor', default='4',
                                 tracking=True)
    towers = fields.Boolean(string='Tower Building', tracking=True)
    no_of_towers = fields.Integer(string='No. of Towers', tracking=True)
    furnishing = fields.Selection([('fully_furnished', 'Fully Furnished'),
                                   ('only_kitchen', 'Only Kitchen Furnished'),
                                   ('only_bed', 'Only BedRoom Furnished'),
                                   ('not_furnished', 'Not Furnished'),
                                   ], string='Furnishing',
                                  default='fully_furnished', tracking=True)
    bed = fields.Integer(string='Bedroom', default=1, tracking=True)
    bathroom = fields.Integer(string='Bathroom', default=1, tracking=True)
    parking = fields.Integer(string='Parking', default=1, tracking=True)
    facing = fields.Selection([('N', 'North(N)'),
                               ('E', 'East(E)'),
                               ('S', 'South(S)'),
                               ('W', 'West(W)'),
                               ('NE', 'North-East(NE)'),
                               ('SE', 'South-East(SE)'),
                               ('SW', 'South-West(SW)'),
                               ('NW', 'North-West(NW)'),
                               ],
                              string='Facing', default='N', tracking=True)
    room_no = fields.Char(string='Flat No./House No.', tracking=True)
    floor = fields.Char(string='Floor', tracking=True)
    total_square_ft = fields.Char(string='Total Square Feet', tracking=True)
    usable_square_ft = fields.Char(string='Usable Square Feet', tracking=True)
    facilities = fields.Text(string='Facilities', tracking=True)

    # Land
    land_name = fields.Char(string='Land Name', tracking=True)
    area_hector = fields.Char(string='Area in Hector', tracking=True)
    land_facilities = fields.Text(string='Land Facility', tracking=True)

    # Commercial
    commercial_name = fields.Char(string='Commercial/Shop Name', tracking=True)
    commercial_type = fields.Selection([('full_commercial', 'Full Commercial'),
                                        ('shops', 'Shops'),
                                        ('big_hall', 'Big Hall')],
                                       string='Commercial Type', tracking=True)
    used_for = fields.Selection([('offices', 'Offices'),
                                 (' retail_stores', ' Retail Stores'),
                                 ('shopping_centres', 'Shopping Centres'),
                                 ('hotels', 'Hotels'),
                                 ('restaurants', 'Restaurants'),
                                 ('pubs', 'Pubs'),
                                 ('cafes', 'Cafes'),
                                 ('sport_facilities', 'Sport Facilities'),
                                 ('medical_centres', 'Medical Centres'),
                                 ('hospitals', 'Hospitals'),
                                 ('nursing_homes', 'Nursing Homes'),
                                 ('other', 'Other Use')
                                 ],
                                string='Used For', tracking=True)
    floor_commercial = fields.Integer(string='Floor Commercial', tracking=True)
    total_floor_commercial = fields.Char(string='Total Floor Commercial',
                                         tracking=True)
    commercial_facilities = fields.Text(string='Commercial Facilities ',
                                        tracking=True)
    other_use = fields.Char(string='Other Use', tracking=True)

    # Industrial
    industry_name = fields.Char(string='Industry Name', tracking=True)
    industry_location = fields.Selection([('inside', 'Inside City'),
                                          ('outside', 'Outside City')],
                                         string='Location', tracking=True)
    industrial_used_for = fields.Selection([('company', 'Company'),
                                            ('warehouses', 'Warehouses'),
                                            ('factories', 'Factories'),
                                            ('other', 'Other')],
                                           string='Usage', tracking=True)
    other_usages = fields.Char(string='Other Usages ', tracking=True)
    industrial_facilities = fields.Text(string='Industrial Facilities  ',
                                        tracking=True)

    # Measurement Details
    room_measurement_ids = fields.One2many('property.room.measurement',
                                           'room_measurement_id',
                                           string='Room Measurement',
                                           tracking=True)
    commercial_measurement_ids = fields.One2many(
        'property.commercial.measurement', 'commercial_measurement_id',
        string='Commercial Measurement', tracking=True)
    industrial_measurement_ids = fields.One2many(
        'property.industrial.measurement', 'industrial_measurement_id',
        string='Industrial Measurement', tracking=True)
    total_room_measure = fields.Integer(
        string='Total Square feets ', compute='_compute_room_measure',
        store=True, tracking=True)
    total_commercial_measure = fields.Integer(
        string='Total Square feet(Commercial)',
        compute='_compute_commercial_measure',
        store=True, tracking=True)
    total_industrial_measure = fields.Integer(
        string='Total Square feet(Industrial)  ',
        compute='_compute_industrial_measure',
        store=True, tracking=True)
    # Smart Button Count
    document_count = fields.Integer(
        string='Document Count', compute='_compute_document_count',
        tracking=True)
    request_count = fields.Integer(
        string='Request Count', compute='_compute_request_count',
        tracking=True)
    booking_count = fields.Monetary(
        string='Booking Count', compute='_compute_booking_count',
        tracking=True)

    contract_count = fields.Integer(
        string='Contract Count', compute='_compute_contract_count',
        tracking=True)

    p_handover_count = fields.Integer(
        string='Handover Count', compute='_compute_p_hanover_count',
        tracking=True)

    p_utility_count = fields.Integer(
        string='Utility Count', compute='_compute_p_utility_count',
        tracking=True)

    p_invoice_count = fields.Integer(
        string='Invoice Count', compute='_compute_p_invoice_count',
        tracking=True)
    lease_contract_count = fields.Integer(
        compute='_compute_lease_contract_count',
        string='Lease Contracts Count',
        tracking=True)
    service_contract_count = fields.Integer(
        compute='_compute_service_contract_count',
        string='Service Contracts Count',
        tracking=True)

    # CRM Lead
    lead_count = fields.Integer(string="Lead Count", compute="_compute_lead",
                                tracking=True)
    lead_opp_count = fields.Integer(
        string="Opportunity Count", compute="_compute_lead", tracking=True)

    # new added field
    plot_no = fields.Char(string='Plot No', tracking=True)
    currency_id = fields.Many2one('res.currency', tracking=True,
                                  string="Rent Currency")
    document_ids = fields.Many2many('documents.document', tracking=True,
                                    string="Document")
    unit_type = fields.Char(string="Unit Type", tracking=True)
    master_bedroom = fields.Integer(string="Master Bedroom", tracking=True)
    build_area = fields.Char("Build-up Area(Sq.M)", tracking=True)
    carpet_area = fields.Char("Carpet Area(Sq.M)", tracking=True)
    property_assets_ids = fields.Many2many('product.product',
                                           'product_property_rel',
                                           'property_id',
                                           string="Property Assets",
                                           tracking=True)
    property_meter_ids = fields.One2many("property.meter", 'property_id',
                                         tracking=True,
                                         string="Property Meter")
    company_id = fields.Many2one('res.company', string="Main Landlord Company",
                                 default=False, tracking=True)
    maintanance_company_id = fields.Many2one('res.company',
                                             string="Maintenance Company",
                                             tracking=True)
    active = fields.Boolean(default=True, tracking=True)
    p_assets_ids = fields.One2many('property.assets', 'property_id',
                                   tracking=True)
    rented_area = fields.Float(string='Rented Area', tracking=True)
    square_area = fields.Selection(
        [('square_meter', 'Square Meter'), ('square_feet', 'Square Feet')],
        string="Square Area", default='square_meter', tracking=True)
    rent_smtr = fields.Monetary(string='Rent/Square Meter', tracking=True)
    pro_company_id = fields.Many2one('res.company',
                                     string="Default Main Landlord Company",
                                     default=lambda self: self.env.company,
                                     tracking=True)
    is_main = fields.Boolean(string="Is Main", compute="_compute_is_main")
    barcode = fields.Binary(string="Lipa QR Code")
    lipa_number = fields.Char(string='Lipa Number', tracking=True, size=10)

    is_any_service_running = fields.Boolean(
        compute='compute_is_any_service_running')

    staff_id = fields.Many2one('res.partner', string='Staff')
    rent = fields.Selection(
        [('per_square_meter', 'Per Square Meter'), ('fixed', 'Fixed')],
        string='Rent Type')

    # @api.model
    # def get_property_details(self):
    #     """For frontdesk get property data"""
    #     properties = self.env['property.details'].sudo().search([])
    #     print('............',properties)
    #     return {'hii':1}

    @api.model
    def _search(self, domain, offset=0, limit=None, order=None):
        context = self._context or {}
        allowed_company_ids = self._context.get('allowed_company_ids')
        if context.get('skip_check'):
            pass
        elif self.env.company.is_maintenance_company:
            domain += [(1, '=', 1)]
        elif allowed_company_ids is None:
            pass
        else:
            domain += [('pro_company_id', 'in', allowed_company_ids)]
        return super(PropertyDetails, self)._search(domain, offset, limit, order)

    @api.depends('contract_history_ids')
    def compute_is_any_service_running(self):
        for rec in self:
            rec.is_any_service_running = False
            contracts = self.env['tenancy.details'].search(
                [('contract_type_name', '=', 'service'),
                 ('contract_type', '=', 'running_contract'),
                 ('property_id', '=', rec.id)], limit=1)
            if contracts:
                rec.is_any_service_running = True

    @api.constrains('lipa_number')
    def _check_required_computed_currencies(self):
        for rec in self:
            if rec.lipa_number:
                if not rec.lipa_number.isdigit():
                    raise ValidationError(
                        "You can not store non digit data as Lipa Number")

    @api.onchange('rented_area', 'rent_smtr')
    def onchange_rented_area_get_tenancy_price(self):
        for rec in self:
            if rec.rent == 'per_square_meter':
                rec.tenancy_price = rec.rented_area * rec.rent_smtr

    @api.onchange('name')
    def onchange_name(self):
        for rec in self:
            exists = self.env['property.details'].search(
                [('name', '=', rec.name)])
            if exists:
                raise UserError(
                    'Property "' + rec.name + '" is already exist in record.Please change the Property Name !!!')

    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        self.ensure_one()
        if default is None:
            default = {}
        if not default.get('name'):
            default.update(name=_('%s (copy)') % (self.name))
        res = super(PropertyDetails, self).copy(default)
        res.stage = 'draft'
        return res

    def _compute_is_main(self):
        if self.env.company.is_maintenance_company:
            self.write({'is_main': True})
        else:
            self.write({'is_main': False})

    def action_property_management_agreement(self):
        for rec in self:
            view = self.env.ref('rental_management.tenancy_details_form_view')
            ctx = {
                'default_property_id': rec.id,
                'default_parent_property_id': rec.parent_property_id.id,
                'default_contract_type_name': 'service'
            }
            return {
                'name': 'Contract',
                'res_model': 'tenancy.details',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'context': ctx,
                'view_id': view.id,
                'target': 'current',
            }

    def open_contract(self):
        for rec in self:
            view = self.env.ref('rental_management.tenancy_details_form_view')
            context = {'default_property_id': rec.id,
                       'default_parent_property_id': rec.parent_property_id.id,
                       'default_rent': rec.rent}
            return {
                'name': 'Contract',
                'res_model': 'tenancy.details',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'view_id': view.id,
                'target': 'current',
                'context': context,
            }

    def action_property_book(self):
        for rec in self:
            rec.write({
                'stage': 'booked',
            })

    def action_property_vacant(self):
        for rec in self:
            rec.write({
                'stage': 'available'
            })

    @api.ondelete(at_uninstall=False)
    def _unlink_(self):
        for rec in self:

            if rec.document_count > 0:
                raise UserError(_(
                    "You can not delete a property! you must first delete it's Documents Record."))

            if rec.request_count > 0:
                raise UserError(_(
                    "You can not delete a property! you must first delete it's Maintenance Record."))

            if rec.booking_count > 0:
                raise UserError(_(
                    "You can not delete a property! you must first delete it's Booking Record."))

            if rec.lead_count > 0:
                raise UserError(_(
                    "You can not delete a property! you must first delete it's Lead or Opportunity Record."))

            # if rec.lead_count > 0:
            #     raise UserError(_(
            #         "You can not delete a property! you must first delete it's Opportunity Record."))

            if rec.lease_contract_count > 0 or rec.service_contract_count > 0:
                raise UserError(_(
                    "You can not delete a property! You have contract records available related to the property."))

            if rec.p_handover_count > 0:
                raise UserError(_(
                    "You can not delete a property! You have property handover records available for the property."))

            if rec.p_utility_count > 0:
                raise UserError(_(
                    "You can not delete a property! You have utility bill records available for the property."))

            if rec.p_invoice_count > 0:
                raise UserError(_(
                    "You can not delete a property! You have invoice records available for the property."))

    @api.onchange('landlord_id', 'maintanance_company_id')
    def _onchange_maintainance_company(self):
        for rec in self:
            if rec.maintanance_company_id:
                rec.landlord_id.maintanance_company_id = rec.maintanance_company_id

    @api.depends('room_measurement_ids')
    def _compute_room_measure(self):
        for rec in self:
            total = 0
            if rec.room_measurement_ids:
                for data in rec.room_measurement_ids:
                    total = total + data.carpet_area
            rec.total_room_measure = total

    @api.depends('sale_lease')
    def _compute_lead(self):
        for rec in self:
            rec.lead_count = self.env['crm.lead'].search_count(
                [('property_id', '=', rec.id)])
            rec.lead_opp_count = self.env['crm.lead'].search_count(
                [('property_id', '=', rec.id), ('type', '=', 'opportunity')])

    @api.depends('commercial_measurement_ids')
    def _compute_commercial_measure(self):
        for rec in self:
            total = 0
            if rec.commercial_measurement_ids:
                for data in rec.commercial_measurement_ids:
                    total = total + data.carpet_area
            rec.total_commercial_measure = total

    @api.depends('industrial_measurement_ids')
    def _compute_industrial_measure(self):
        for rec in self:
            total = 0
            if rec.industrial_measurement_ids:
                for data in rec.industrial_measurement_ids:
                    total = total + data.carpet_area
            rec.total_industrial_measure = total

    @api.depends('extra_service_ids')
    def _compute_extra_service_cost(self):
        for rec in self:
            amount = 0.0
            if rec.extra_service_ids:
                for data in rec.extra_service_ids:
                    amount = amount + data.price
            rec.extra_service_cost = amount

    # Sequence Create
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('property_seq', 'New') == 'New':
                vals['property_seq'] = self.env['ir.sequence'].next_by_code(
                    'property.details') or 'New'

        res = super(PropertyDetails, self).create(vals_list)
        return res

    def _compute_display_name(self):
        for rec in self:
            type_label = dict(rec._fields['type'].selection).get(rec.type, '')
            if rec.is_parent_property and rec.parent_property_id:
                rec.display_name = '%s - %s - %s' % (rec.name, rec.parent_property_id.name, type_label)
            else:
                rec.display_name = '%s - %s' % (rec.name, type_label)

    # Buttons
    def action_in_available(self):
        for rec in self:
            rec.stage = 'available'

    def action_occupied_by_staff(self):
        for rec in self:
            rec.stage = 'occupied_by_staff'

    def action_in_sold(self):
        if self.sale_lease == 'for_sale':
            self.stage = 'sold'
        else:
            message = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'info',
                    'title': 'You need to set "Price/Rent" to "For Sale" to procced',
                    'sticky': False,
                }
            }
            return message

    # Smart Button
    def _compute_document_count(self):
        for rec in self:
            document_count = self.env['property.documents'].search_count(
                [('property_id', '=', rec.id)])
            rec.document_count = document_count

    def _compute_lease_contract_count(self):
        companies = self.env['res.company'].browse(
            self._context.get('allowed_company_ids')).ids
        for rec in self:
            companies.append(rec.pro_company_id.id)
            lease_contract_count = self.env[
                'tenancy.details'].sudo().search_count(
                [('property_id', '=', rec.id),
                 ('pro_company_id', 'in', companies),
                 ('contract_type_name', '=', 'lease')])
            rec.lease_contract_count = lease_contract_count

    def _compute_service_contract_count(self):
        companies = self.env['res.company'].browse(
            self._context.get('allowed_company_ids')).ids
        for rec in self:
            companies.append(rec.pro_company_id.id)
            service_contract_count = self.env[
                'tenancy.details'].sudo().search_count(
                [('property_id', '=', rec.id),
                 ('pro_company_id', 'in', companies),
                 ('contract_type_name', '=', 'service')])
            rec.service_contract_count = service_contract_count

    def _compute_booking_count(self):
        for rec in self:
            count = self.sold_booking_id.book_price
            rec.booking_count = count

    def _compute_request_count(self):
        for rec in self:
            request_count = self.env['maintenance.request'].search_count(
                [('property_id', '=', rec.id)])
            rec.request_count = request_count

    def _compute_contract_count(self):
        for rec in self:
            contract_count = self.env['tenancy.details'].search_count(
                [('property_id', '=', rec.id),
                 ('pro_company_id', '=', self.env.company.id)])
            rec.contract_count = contract_count

    def _compute_p_hanover_count(self):
        for rec in self:
            p_handover_count = self.env['handover.property'].search_count(
                [('property_id', '=', rec.id)])
            rec.p_handover_count = p_handover_count

    def _compute_p_utility_count(self):
        for rec in self:
            p_utility_count = self.env['utility.bill'].search_count(
                [('property_id', '=', rec.id)])
            rec.p_utility_count = p_utility_count

    def _compute_p_invoice_count(self):
        for rec in self:
            p_invoice_count = self.env['account.move'].search_count(
                [('tenancy_property_id', '=', rec.id)])
            rec.p_invoice_count = p_invoice_count

    def action_maintenance_request(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Request',
            'res_model': 'maintenance.request',
            'domain': [('property_id', '=', self.id)],
            'context': {'default_property_id': self.id},
            'view_mode': 'kanban,list',
            'target': 'current'
        }

    def action_contarct(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Contract',
            'res_model': 'tenancy.details',
            'domain': [('property_id', '=', self.id),
                       ('pro_company_id', '=', self.env.company.id)],
            'context': {'default_property_id': self.id,
                        'default_parent_property_id': self.parent_property_id.id, },
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }

    def action_lease_contracts(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Lease Contracts',
            'res_model': 'tenancy.details',
            'domain': [('property_id', '=', self.id), (
                'company_id', 'in', self.env['res.company'].browse(
                    self._context.get('allowed_company_ids')).ids),
                       ('contract_type_name', '=', 'lease')],
            'context': {'default_property_id': self.id,
                        'default_contract_type_name': 'lease',
                        'default_parent_property_id': self.parent_property_id.id,
                        'default_rent': self.rent},
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }

    def action_property_management_contracts(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Property Management Contract',
            'res_model': 'tenancy.details',
            'domain': [('property_id', '=', self.id), (
                'company_id', 'in', self.env['res.company'].browse(
                    self._context.get('allowed_company_ids')).ids),
                       ('contract_type_name', '=', 'service')],
            'context': {'default_property_id': self.id,
                        'default_contract_type_name': 'service',
                        'default_parent_property_id': self.parent_property_id.id,
                        'default_rent': self.rent},
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }

    def action_p_handover(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Handover',
            'res_model': 'handover.property',
            'domain': [('property_id', '=', self.id)],
            'context': {'default_property_id': self.id},
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }

    def action_p_utility_bills(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Utility Bills',
            'res_model': 'utility.bill',
            'domain': [('property_id', '=', self.id)],
            'context': {'default_property_id': self.id},
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }

    def action_p_invoice(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Invoice',
            'res_model': 'account.move',
            'domain': [('tenancy_property_id', '=', self.id)],
            'context': {'default_tenancy_property_id': self.id},
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }

    def action_property_document(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Document',
            'res_model': 'property.documents',
            'domain': [('property_id', '=', self.id)],
            'context': {'default_property_id': self.id},
            'view_mode': 'list',
            'target': 'current'
        }

    def action_sale_booking(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Booking Information',
            'res_model': 'property.vendor',
            'domain': [('property_id', '=', self.id)],
            'context': {'default_property_id': self.id},
            'view_mode': 'list',
            'target': 'current'
        }

    def action_crm_lead(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Leads',
            'res_model': 'crm.lead',
            'domain': [('property_id', '=', self.id)],
            'context': {'default_property_id': self.id},
            'view_mode': 'list,form',
            'target': 'current'
        }

    def action_crm_lead_opp(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Opportunity',
            'res_model': 'crm.lead',
            'domain': [('property_id', '=', self.id),
                       ('type', '=', 'opportunity')],
            'context': {'default_property_id': self.id,
                        'default_type': 'opportunity'},
            'view_mode': 'list,form',
            'target': 'current'
        }

    @api.onchange('is_parent_property', 'parent_property_id')
    def _onchange_parent_property_type(self):
        for rec in self:
            if not rec.is_parent_property and not rec.parent_property_id:
                return
            rec.type = rec.parent_property_id.type
            rec.residence_type = rec.parent_property_id.residence_type
            rec.total_floor = rec.parent_property_id.total_floor
            rec.towers = rec.parent_property_id.towers
            rec.no_of_towers = rec.parent_property_id.no_of_towers
            rec.industry_location = rec.parent_property_id.industry_location
            rec.commercial_type = rec.parent_property_id.commercial_type
            rec.plot_no = rec.parent_property_id.plot_no

    @api.model
    def property_status_contract_based(self):
        records = self.env['property.details'].sudo().search([])
        for rec in records:
            lease_contracts = self.env['tenancy.details'].sudo().search(
                [('contract_type_name', '=', 'lease'),
                 ('property_id', '=', rec.id)])
            if not lease_contracts and rec.sale_lease == 'for_tenancy':
                rec.write({
                    'stage': 'available',
                })
            running_lease_contracts = self.env[
                'tenancy.details'].sudo().search(
                [('contract_type_name', '=', 'lease'),
                 ('contract_type', '=', 'running_contract'),
                 ('property_id', '=', rec.id)], limit=1)
            if not running_lease_contracts and rec.sale_lease == 'for_tenancy':
                rec.write({
                    'stage': 'available',
                })

    @api.model
    def fix_property_status(self):
        records = self.env['property.details'].sudo().search([])
        for rec in records:
            running_lease_contracts = self.env[
                'tenancy.details'].sudo().search(
                [('contract_type_name', '=', 'lease'),
                 ('property_id', '=', rec.id),
                 ('contract_type', '=', 'running_contract')], limit=1)
            if running_lease_contracts and rec.sale_lease == 'for_tenancy':
                rec.write({
                    'stage': 'on_lease'
                })

    @api.model
    def fix_rent_type(self):
        records = self.env['property.details'].sudo().search([])
        for rec in records:
            if rec.tenancy_price > 0:
                rec.rent = 'fixed'
            if rec.rent_smtr > 0:
                rec.rent = 'per_square_meter'

    @api.model
    def get_property_stats(self, company_ids):

        # Property Stages
        property = self.env['property.details']
        avail_property = property.sudo().search_count(
            ['&', ('stage', '=', 'available'),
             ('pro_company_id', 'in', company_ids)])
        lease_property = property.sudo().search_count(
            ['&', ('stage', '=', 'on_lease'),
             ('pro_company_id', 'in', company_ids)])
        sold_property = property.sudo().search_count(
            ['&', ('stage', '=', 'sold'),
             ('pro_company_id', 'in', company_ids)])

        for rec in company_ids:
            company_id = self.env['res.company'].browse(rec)

            if company_id.is_maintenance_company:
                get_total_property = property.sudo().read_group([], fields=[],
                                                                groupby=[])
                count = get_total_property[0].get('__count')
                total_property = count
            else:
                get_total_property = property.sudo().read_group(
                    [('pro_company_id', 'in', company_ids)], fields=[],
                    groupby=[])
                count = get_total_property[0].get('__count')
                total_property = count

        currency_symbol = self.env['res.currency'].search([]).mapped('symbol')
        draft_contract = self.env['tenancy.details'].sudo().search_count(
            ['&', ('contract_type', '=', 'new_contract'),
             ('company_id', 'in', company_ids)])
        running_contract = self.env['tenancy.details'].sudo().search_count(
            ['&', ('contract_type', '=', 'running_contract'),
             ('company_id', 'in', company_ids)])
        expire_contract = self.env['tenancy.details'].sudo().search_count(
            ['&', ('contract_type', '=', 'expire_contract'),
             ('company_id', 'in', company_ids)])
        sale_sold = self.env['property.vendor'].sudo(
        ).search_count(
            ['&', ('stage', '=', 'sold'), ('company_id', 'in', company_ids)])

        # Total Tenancy and Sold Information
        sold_total = self.env['property.vendor'].search(
            ['&', ('stage', '=', 'sold'),
             ('company_id', 'in', company_ids)]).mapped('sale_price')
        total = 0
        for price in sold_total:
            total = total + price
        total_str = str(total)
        full_tenancy_total = self.env['rent.invoice'].search(
            ['|', '&', ('type', '=', 'rent'), ('type', '=', 'full_rent'),
             ('company_id', 'in', company_ids)])
        final_rent = 0
        for rent in full_tenancy_total:
            final_rent = final_rent + rent.rent_invoice_id.amount_total
        final_rent_str = str(final_rent)

        # Pending Invoice
        pending_invoice = self.env['rent.invoice'].search_count(
            ['&', ('payment_state', '=', 'not_paid'),
             ('company_id', 'in', company_ids)])

        # Property Type
        land_property = property.sudo().search_count([('type', '=', 'land')])
        residential_property = property.sudo().search_count(
            ['&', ('type', '=', 'residential'),
             ('pro_company_id', 'in', company_ids)])
        commercial_property = property.sudo().search_count(
            ['&', ('type', '=', 'commercial'),
             ('pro_company_id', 'in', company_ids)])
        industrial_property = property.sudo().search_count(
            ['&', ('type', '=', 'industrial'),
             ('pro_company_id', 'in', company_ids)])
        property_type = [['Land', 'Residential', 'Commercial', 'Industrial'],
                         [land_property, residential_property,
                          commercial_property, industrial_property]]
        property_stage = [
            ['Available Properties', 'Sold Properties', 'On Lease'],
            [avail_property, sold_property, lease_property]]

        data = {
            'avail_property': avail_property,
            'lease_property': lease_property,
            'sold_property': sold_property,
            'total_property': total_property,
            'sold_total': total_str + ' ' + currency_symbol[
                0] if currency_symbol else "",
            'rent_total': final_rent_str + ' ' + currency_symbol[
                0] if currency_symbol else "",
            'draft_contract': draft_contract,
            'running_contract': running_contract,
            'expire_contract': expire_contract,
            'sale_sold': sale_sold,
            'property_type': property_type,
            'property_stage': property_stage,
            'pending_invoice': pending_invoice,
            'tenancy_top_broker': self.get_top_broker(),
            'due_paid_amount': self.due_paid_amount(),
        }
        return data

    def get_top_broker(self):
        broker_tenancy = {}
        broker_sold = {}
        for group in self.env['tenancy.details'].read_group(
                [('is_any_broker', '=', True)],
                ['broker_id'],
                ['broker_id'], limit=5):
            if group['broker_id']:
                name = self.env['res.partner'].sudo().browse(
                    int(group['broker_id'][0])).name
                broker_tenancy[name] = group['broker_id_count']
        for group in self.env['property.vendor'].read_group(
                [('is_any_broker', '=', True), ('stage', '=', 'sold')],
                ['broker_id'],
                ['broker_id'], limit=5):
            if group['broker_id']:
                name = self.env['res.partner'].sudo().browse(
                    int(group['broker_id'][0])).name
                broker_sold[name] = group['broker_id_count']

        brokers_tenancy_list = dict(
            sorted(broker_tenancy.items(), key=lambda x: x[1], reverse=True))
        broker_sold_list = dict(
            sorted(broker_sold.items(), key=lambda x: x[1], reverse=True))
        return [list(brokers_tenancy_list.keys()),
                list(brokers_tenancy_list.values()),
                list(broker_sold_list.keys()),
                list(broker_sold_list.values())]

    def due_paid_amount(self):
        sold = {}
        tenancy = {}
        not_paid_amount_sold = 0.0
        paid_amount_sold = 0.0
        not_paid_amount_tenancy = 0.0
        paid_amount_tenancy = 0.0
        property_sold = self.env['account.move'].sudo().search(
            [('sold_id', '!=', False)])
        for data in property_sold:
            if data.sold_id.stage == "sold":
                if data.payment_state == "not_paid":
                    not_paid_amount_sold = not_paid_amount_sold + data.amount_total
                if data.payment_state == "paid":
                    paid_amount_sold = paid_amount_sold + data.amount_total
        sold['Property Sold Due'] = not_paid_amount_sold
        sold['Property Sold Paid'] = paid_amount_sold
        property_tenancy = self.env['rent.invoice'].sudo().search([])
        for rec in property_tenancy:
            if rec.payment_state == 'not_paid':
                not_paid_amount_tenancy = not_paid_amount_tenancy + \
                                          rec.rent_invoice_id.amount_total
            if rec.payment_state == 'paid':
                paid_amount_tenancy = paid_amount_tenancy + rec.rent_invoice_id.amount_total
        tenancy['Tenancy Due'] = not_paid_amount_tenancy
        tenancy['Tenancy Paid'] = paid_amount_tenancy
        return [list(sold.keys()), list(sold.values()), list(tenancy.keys()),
                list(tenancy.values())]

    def action_gmap_location(self):
        if self.longitude and self.latitude:
            longitude = self.longitude
            latitude = self.latitude
            http_url = 'https://maps.google.com/maps?q=loc:' + latitude + ',' + longitude
            return {
                'type': 'ir.actions.act_url',
                'target': 'new',
                'url': http_url,
            }
        else:
            raise ValidationError(
                "! Enter Proper Longitude and Latitude Values")

    @api.onchange('is_parent_property', 'parent_property_id')
    def _onchange_parent_landlord(self):
        for rec in self:
            if rec.is_parent_property and rec.parent_property_id:
                rec.landlord_id = rec.parent_property_id.landlord_id.id
                rec.zip = rec.parent_property_id.zip
                rec.street = rec.parent_property_id.street
                rec.street2 = rec.parent_property_id.street2
                rec.city_id = rec.parent_property_id.city_id.id
                rec.country_id = rec.parent_property_id.country_id.id
                rec.state_id = rec.parent_property_id.state_id.id
                rec.website = rec.parent_property_id.website


# Property Meter
class PropertyMeter(models.Model):
    _name = 'property.meter'
    _description = 'Property Meter Details'

    serial_no = fields.Integer(string="Sr.No", compute="_compute_sr_no")
    meter_no = fields.Integer(string="Meter No")
    meter_type_id = fields.Many2one('meter.type', string="Meter Type")
    installation_date = fields.Date(string="Date", default=fields.Date.today)
    initial_reading = fields.Integer(string="Initial Reading")
    is_status = fields.Boolean(string="Status")
    property_id = fields.Many2one('property.details', string="Property")

    def _compute_sr_no(self):
        no = 0
        for line in self:
            no += 1
            line.serial_no = no


# Room Measurement
class PropertyRoomMeasurement(models.Model):
    _name = 'property.room.measurement'
    _description = 'Room Property Measurement Details'

    type_room = fields.Selection([('hall', 'Hall'),
                                  ('bed_room', 'Bed Room'),
                                  ('kitchen', 'Kitchen'),
                                  ('drawing_room', 'Drawing Room'),
                                  ('bathroom', 'Bathroom'),
                                  ('store_room', 'Store Room'),
                                  ('balcony', 'Balcony'),
                                  ('wash_area', 'Wash Area'),
                                  ],
                                 string='House Section')
    length = fields.Integer(string='Length(ft)')
    width = fields.Integer(string='Width(ft)')
    height = fields.Integer(string='Height(ft)')
    no_of_unit = fields.Integer(string="No of Unit", default=1)
    carpet_area = fields.Integer(
        string='Carpet Area(ft²)', compute='_compute_carpet_area')
    measure = fields.Char(string='ft²', default='ft²', readonly=True)
    room_measurement_id = fields.Many2one(
        'property.details', string='Room Details')

    @api.depends('length', 'width')
    def _compute_carpet_area(self):
        for rec in self:
            total = 0
            if rec.length and rec.width:
                total = rec.length * rec.width * rec.no_of_unit
            rec.carpet_area = total


class PropertyCommercialMeasurement(models.Model):
    _name = 'property.commercial.measurement'
    _description = 'Commercial Property Measurement Details'

    shops = fields.Char(string='Section')
    length = fields.Integer(string='Length(ft)')
    width = fields.Integer(string='Width(ft)')
    height = fields.Integer(string='Height(ft)')
    carpet_area = fields.Integer(
        string='Area(ft²)', compute='_compute_carpet_area')
    measure = fields.Char(string='ft²', default='ft²', readonly=True)
    commercial_measurement_id = fields.Many2one(
        'property.details', string='Commercial Details')
    no_of_unit = fields.Integer(string="No of Unit", default=1)

    @api.depends('length', 'width')
    def _compute_carpet_area(self):
        for rec in self:
            total = 0
            if rec.length and rec.width:
                total = rec.length * rec.width * rec.no_of_unit
            rec.carpet_area = total


class PropertyIndustrialMeasurement(models.Model):
    _name = 'property.industrial.measurement'
    _description = 'Industrial Property Measurement Details'

    asset = fields.Char(string='industrial Asset')
    length = fields.Integer(string='Length(ft)')
    width = fields.Integer(string='Width(ft)')
    height = fields.Integer(string='Height(ft)')
    carpet_area = fields.Integer(
        string='Area(ft²)', compute='_compute_carpet_area')
    measure = fields.Char(string='ft²', default='ft²', readonly=True)
    industrial_measurement_id = fields.Many2one(
        'property.details', string='Industrial Details')
    no_of_unit = fields.Integer(string="No of Unit", default=1)

    @api.depends('length', 'width')
    def _compute_carpet_area(self):
        for rec in self:
            total = 0
            if rec.length and rec.width:
                total = rec.length * rec.width * rec.no_of_unit
            rec.carpet_area = total


class DocumentsDocument(models.Model):
    _inherit = 'documents.document'

    property_id = fields.Many2one(
        'property.details', string='Property Name', readonly=True)


class PropertyDocuments(models.Model):
    _name = 'property.documents'
    _description = 'Document related to Property'
    _rec_name = 'doc_type'

    property_id = fields.Many2one(
        'property.details', string='Property Name', readonly=True)
    document_date = fields.Date(string='Date', default=fields.Date.today())
    doc_type = fields.Selection([('photos', 'Photo'),
                                 ('brochure', 'Brochure'),
                                 ('certificate', 'Certificate'),
                                 ('insurance_certificate',
                                  'Insurance Certificate'),
                                 ('utilities_insurance',
                                  'Utilities Certificate')],
                                string='Document Type', required=True)
    document = fields.Binary(string='Documents', required=True)
    file_name = fields.Char(string='File Name')
    tenancy_id = fields.Many2one(
        'tenancy.details', string='Contract', readonly=True)


class PropertyAmenities(models.Model):
    _name = 'property.amenities'
    _description = 'Details About Property Amenities'
    _rec_name = 'title'

    image = fields.Binary(string='Image')
    title = fields.Char(string='Title')


class PropertySpecification(models.Model):
    _name = 'property.specification'
    _description = 'Details About Property Specification'
    _rec_name = 'title'

    image = fields.Image(string='Image')
    title = fields.Char(string='Title')
    description_line1 = fields.Char(string='Description')
    description_line2 = fields.Char(string='Description Line 2')
    description_line3 = fields.Char(string='Description Line 3')


class CertificateType(models.Model):
    _name = 'certificate.type'
    _description = 'Type Of Certificate'
    _rec_name = 'type'

    type = fields.Char(string='Type')


class PropertyCertificate(models.Model):
    _name = 'property.certificate'
    _description = 'Property Related All Certificate'
    _rec_name = 'type_id'

    type_id = fields.Many2one('certificate.type', string='Type')
    expiry_date = fields.Date(string='Expiry Date')
    responsible = fields.Char(string='Responsible')
    note = fields.Char(string='Note')
    property_id = fields.Many2one('property.details', string='Property')


class ParentProperty(models.Model):
    _name = 'parent.property'
    _description = 'Parent Property Details'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(string='Name')
    image = fields.Binary(string='Image')
    amenities_ids = fields.Many2many('property.amenities', string='Amenities')
    property_specification_ids = fields.Many2many(
        'property.specification', string='Specification')
    zip = fields.Char(string='Pin Code', size=6)
    street = fields.Char(string='Street1')
    street2 = fields.Char(string='Street2')
    city = fields.Char(string='City ')
    city_id = fields.Many2one('property.res.city', string='City')
    country_id = fields.Many2one('res.country', 'Country')
    state_id = fields.Many2one(
        "res.country.state", string='State', readonly=False, store=True,
        domain="[('country_id', '=?', country_id)]")
    landlord_id = fields.Many2one('res.partner', string='Landlord',
                                  domain=lambda self: [
                                      ('is_landlord', '=', True)])

    website = fields.Char(string='Website')
    airport = fields.Char(string='Airport')
    national_highway = fields.Char(string='National Highway')
    metro_station = fields.Char(string='Metro Station')
    metro_city = fields.Char(string='Metro City')
    school = fields.Char(string="School")
    hospital = fields.Char(string="Hospital")
    shopping_mall = fields.Char(string="Mall")
    park = fields.Char(string="Park")
    nearby_connectivity_ids = fields.Many2many(
        'property.connectivity', string="Nearby Connectivity ")
    type = fields.Selection(
        [('residential', 'Residential'), ('commercial',
                                          'Commercial'),
        ('industrial', 'Industrial')],
        string='Property Type', default="residential")
    property_count = fields.Integer(
        string="Property Count", compute="_compute_properties")
    utility_rate_ids = fields.One2many("meter.type.lines",
                                       inverse_name="main_property_id")

    # Residential
    residence_type = fields.Selection([('apartment', 'Apartment'),
                                       ('bungalow', 'Bungalow'),
                                       ('vila', 'Vila'),
                                       ('raw_house', 'Raw House'),
                                       ('duplex', 'Duplex House'),
                                       ('single_studio', 'Single Studio')],
                                      string='Type of Residence')
    total_floor = fields.Integer(string='Total Floor', default='4')
    towers = fields.Boolean(string='Tower Building')
    no_of_towers = fields.Integer(string='No. of Towers')

    # Commercial
    commercial_type = fields.Selection([('full_commercial', 'Full Commercial'),
                                        ('shops', 'Shops'),
                                        ('big_hall', 'Big Hall')],
                                       string='Commercial Type')

    # Industrial
    industry_location = fields.Selection([('inside', 'Inside City'),
                                          ('outside', 'Outside City')],
                                         string='Location')

    plot_no = fields.Char(string='Plot No')
    document_ids = fields.Many2many('documents.document',
                                    'property_document_default_rel',
                                    'property_id', 'document_id')
    property_images_ids = fields.One2many('property.images',
                                          'parent_property_id',
                                          string='Images')
    active = fields.Boolean(default=True)
    connectivity_ids = fields.One2many(
        'property.connectivity.line', 'property_id',
        string="Nearby Connectivity")

    company_id = fields.Many2one(
        'res.company', string='Company', default=lambda self: self.env.company)

    maintenance_incharge_id = fields.Many2one('res.partner',
                                              string='Maintenance Incharge')

    sold_property_count = fields.Integer(
        compute='_compute_sold_property_count')
    booked_property_count = fields.Integer(
        compute="_compute_booked_property_count")
    on_lease_property_count = fields.Integer(
        compute="_compute_on_lease_property_count")
    available_property_count = fields.Integer(
        compute="_compute_available_property_count")

    @api.ondelete(at_uninstall=False)
    def _unlink_(self):
        for rec in self:
            if rec.property_count > 0:
                raise UserError(_(
                    "You can not delete a main property! you must first delete it's child property."))

            if rec.document_ids:
                raise UserError(_(
                    "You can not delete a main property! you must first delete documents."))

            if rec.property_images_ids:
                raise UserError(_(
                    "You can not delete a main property! you must first delete Images."))

            if rec.amenities_ids:
                raise UserError(_(
                    "You can not delete a main property! you must first delete Amenities."))

            if rec.property_specification_ids:
                raise UserError(_(
                    "You can not delete a main property! you must first delete Facilities."))

    @api.depends('sold_property_count')
    def _compute_sold_property_count(self):
        for rec in self:
            sold_properties = self.env['property.details'].sudo().search_count(
                [('stage', '=', 'sold'), ('parent_property_id', '=', rec.id)])
            rec.sold_property_count = sold_properties

    @api.depends('booked_property_count')
    def _compute_booked_property_count(self):
        for rec in self:
            booked_properties = self.env[
                'property.details'].sudo().search_count(
                [('stage', '=', 'booked'),
                 ('parent_property_id', '=', rec.id)])
            rec.booked_property_count = booked_properties

    @api.depends('on_lease_property_count')
    def _compute_on_lease_property_count(self):
        for rec in self:
            on_lease_properties = self.env[
                'property.details'].sudo().search_count(
                [('stage', '=', 'on_lease'),
                 ('parent_property_id', '=', rec.id)])
            rec.on_lease_property_count = on_lease_properties

    @api.depends('available_property_count')
    def _compute_available_property_count(self):
        for rec in self:
            available_properties = self.env[
                'property.details'].sudo().search_count(
                [('stage', '=', 'available'),
                 ('parent_property_id', '=', rec.id)])
            rec.available_property_count = available_properties

    @api.model
    def send_property_report_mail(self):
        # For excel
        workbook = openpyxl.Workbook()

        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

        sheet1 = workbook.active
        sheet1.title = "Parent Property Wise Occupancy"
        sheet1.merge_cells('A1:G1')
        h1 = sheet1['A1']
        h1.value = 'Parent Property Wise Occupancy'
        h1.font = bold_font
        h1.alignment = center_align
        h1.fill = orange_fill

        sh1_headers = ["Parent Property", "Total Properties", "Sold", "Booked", "On Lease", "Available", "Occupancy %"]
        for col, header in enumerate(sh1_headers, 1):
            cell = sheet1.cell(row=2, column=col, value=header)
            cell.font = bold_font
            cell.fill = orange_fill
            cell.alignment = left_align if col == 1 else right_align

        for col_idx, width in enumerate([20, 18, 10, 10, 10, 10, 12], 1):
            sheet1.column_dimensions[get_column_letter(col_idx)].width = width

        sheet1_row = 3
        total_properties = 0
        total_sold_properties = 0
        total_booked_properties = 0
        total_on_lease_properties = 0
        total_available_properties = 0
        # Excel Sheet2
        sheet2 = workbook.create_sheet("Rent Collection and Outstanding")
        sheet2.merge_cells('A1:G1')
        h2 = sheet2['A1']
        h2.value = 'Parent Property Wise Rent Collection and Outstanding'
        h2.font = bold_font
        h2.alignment = center_align
        h2.fill = yellow_fill
        sh2_headers = ["Parent Property", "YTD (Expected)", "YTD (Collected)", "YTD (Outstanding)",
                        "Last Month (Expected)", "Last Month (Collected)", "Last Month (Outstanding)"]
        for col, header in enumerate(sh2_headers, 1):
             cell = sheet2.cell(row=2, column=col, value=header)
             cell.font = bold_font
             cell.fill = yellow_fill
             cell.alignment = left_align if col == 1 else right_align
        for col_idx, width in enumerate([20, 16, 16, 18, 22, 22, 22], 1):
             sheet2.column_dimensions[get_column_letter(col_idx)].width = width
        sheet2_row = 3
        # Excel Sheet3
        sheet3 = workbook.create_sheet("Outstanding Ageing")
        sheet3.merge_cells('A1:G1')
        h3 = sheet3['A1']
        h3.value = 'Outstanding Ageing'
        h3.font = bold_font
        h3.alignment = center_align
        h3.fill = blue_fill
        sh3_headers = ["Tenant/Customer", "Property", "1-30 Days", "31-60 Days", "61-90 Days", "91-120 Days", "> 120 Days"]
        for col, header in enumerate(sh3_headers, 1):
             cell = sheet3.cell(row=2, column=col, value=header)
             cell.font = bold_font
             cell.fill = blue_fill
             cell.alignment = left_align if col <= 2 else right_align
        for col_idx, width in enumerate([22, 22, 12, 12, 12, 12, 12], 1):
             sheet3.column_dimensions[get_column_letter(col_idx)].width = width
        sheet3_row = 3
        # For mail
        mail_template = self.env.ref(
            "rental_management.property_wise_occupancy_report_mail").sudo()
        parent_properties = self.env['parent.property'].sudo().search([])
        sub_properties = self.env['property.details'].with_context(skip_check=True).sudo().search(
            [])
        invoice = self.env['rent.invoice'].sudo()
        company_ids = self.env['res.company'].sudo().search([]).ids
        collection_data = []
        ageing_data = []
        property_data = []
        today = fields.Date.today()
        first_day_current_month = datetime.date(today.year, today.month, 1)
        last_day_previous_month = first_day_current_month - datetime.timedelta(
            days=1)
        total_ytd_expected = 0
        total_ytd_collected = 0
        total_ytd_outstanding = 0
        total_prev_expected = 0
        total_prev_collected = 0
        total_prev_outstanding = 0

        for rec in parent_properties:
            # Mail
            t_props = self.env['property.details'].with_context(
                skip_check=True).sudo().search_count(
                [('parent_property_id', '=', rec.id), ('pro_company_id', 'in', company_ids)])
            t_sold_props = self.env['property.details'].with_context(
                skip_check=True).sudo().search_count(
                [('parent_property_id', '=', rec.id), ('stage', '=', 'sold'),
                 ('pro_company_id', 'in', company_ids)])
            t_booked_props = self.env['property.details'].with_context(
                skip_check=True).sudo().search_count(
                [('parent_property_id', '=', rec.id), ('stage', '=', 'booked'),
                 ('pro_company_id', 'in', company_ids)])
            t_lease_props = self.env['property.details'].with_context(
                skip_check=True).sudo().search_count(
                [('parent_property_id', '=', rec.id), ('stage', '=', 'on_lease'),
                 ('pro_company_id', 'in', company_ids)])
            t_available_props = self.env['property.details'].with_context(
                skip_check=True).sudo().search_count(
                [('parent_property_id', '=', rec.id), ('stage', '=', 'available'),
                 ('pro_company_id', 'in', company_ids)])
            data = {
                'name': rec.name,
                'property_count': t_props,
                'sold_property_count': t_sold_props,
                'booked_property_count': t_booked_props,
                'on_lease_property_count': t_lease_props,
                'available_property_count': t_available_props,
            }
            property_data.append(data)
            # Excel Sheet1
            occupancy_percentage = 0
            if t_props - t_booked_props - t_sold_props > 0:
                occupancy_percentage = round(
                    (t_lease_props * 100) / (t_props - t_booked_props - t_sold_props), 2)
            for col_idx, val in enumerate([rec.name, t_props, t_sold_props, t_booked_props, t_lease_props, t_available_props, f"{occupancy_percentage} %"], 1):
                cell = sheet1.cell(row=sheet1_row, column=col_idx, value=val)
                cell.alignment = left_align if col_idx == 1 else right_align
            total_properties += t_props
            total_sold_properties += t_sold_props
            total_booked_properties += t_booked_props
            total_on_lease_properties += t_lease_props
            total_available_properties += t_available_props
            sheet1_row += 1

            # Mail
            child_properties = self.env['property.details'].with_context(
                skip_check=True).sudo().search(
                [('parent_property_id', '=', rec.id)]).ids
            ytd_expected = 0
            ytd_collected = 0
            ytd_outstanding = 0
            prev_month_expected = 0
            prev_month_collected = 0
            prev_month_outstanding = 0

            invoice_records = invoice.sudo().search(
                [('property_id', 'in', child_properties)])
            for record in invoice_records:
                amount = record.amount
                invoice_amount = record.rent_invoice_id.amount_total_in_currency_signed
                invoice_remaining_amount = record.rent_invoice_id.amount_residual
                us_currency = self.env['res.currency'].sudo().search(
                    [('name', '=', 'USD')])
                if record.currency_id.name != 'USD':
                    current_currency = self.env['res.currency'].sudo().search(
                        [('id', '=', record.currency_id.id)])
                    convert_currency = current_currency._convert(amount,
                                                                 us_currency,
                                                                 self.env.company,
                                                                 record.invoice_date)
                    amount = convert_currency
                if record.rent_invoice_id.currency_id.name != 'USD':
                    invoice_current_currency = self.env[
                        'res.currency'].sudo().search(
                        [('id', '=', record.rent_invoice_id.currency_id.id)])
                    convert_invoice_currency = invoice_current_currency._convert(
                        invoice_amount, us_currency,
                        self.env.company,
                        record.rent_invoice_id.invoice_date)
                    covert_remaining_amount_currency = invoice_current_currency._convert(
                        invoice_remaining_amount,
                        us_currency, self.env.company,
                        record.rent_invoice_id.invoice_date)
                    invoice_amount = convert_invoice_currency
                    invoice_remaining_amount = covert_remaining_amount_currency
                if record.tenancy_id.contract_type_name == 'lease' and record.invoice_date.year == today.year and record.invoice_date <= today:
                    ytd_expected += amount
                    if record.payment_state in ['paid',
                                                'in_payment']:
                        ytd_collected += invoice_amount
                    if record.payment_state == 'not_paid':
                        ytd_outstanding += invoice_remaining_amount
                    if record.payment_state == 'partial':
                        ytd_collected += (
                                invoice_amount - invoice_remaining_amount)
                        ytd_outstanding += invoice_remaining_amount
                if record.tenancy_id.contract_type_name == 'lease' and record.invoice_date.month == last_day_previous_month.month and record.invoice_date.year == last_day_previous_month.year:
                    prev_month_expected += amount
                    if record.payment_state in ['paid', 'in_payment']:
                        prev_month_collected += invoice_amount
                    if record.payment_state == 'not_paid':
                        prev_month_outstanding += invoice_remaining_amount
                    if record.payment_state == 'partial':
                        prev_month_collected += (
                                invoice_amount - invoice_remaining_amount)
                        prev_month_outstanding += invoice_remaining_amount
            total_ytd_expected += ytd_expected
            total_ytd_collected += ytd_collected
            total_ytd_outstanding += ytd_outstanding
            total_prev_expected += prev_month_expected
            total_prev_collected += prev_month_collected
            total_prev_outstanding += prev_month_outstanding
            # Excel Sheet 2
            for col_idx, val in enumerate([rec.name, round(ytd_expected, 2), round(ytd_collected, 2), round(ytd_outstanding, 2), round(prev_month_expected, 2), round(prev_month_collected, 2), round(prev_month_outstanding, 2)], 1):
                cell = sheet2.cell(row=sheet2_row, column=col_idx, value=val)
                cell.alignment = left_align if col_idx == 1 else right_align
            sheet2_row += 1
            c_data = {
                'parent_property': rec.name,
                'ytd_expected': ytd_expected,
                'ytd_collected': ytd_collected,
                'ytd_outstanding': ytd_outstanding,
                'prev_month_expected': prev_month_expected,
                'prev_month_collected': prev_month_collected,
                'prev_month_outstanding': prev_month_outstanding,
            }
            collection_data.append(c_data)
        # Excel Sheet1
        total_occupancy_percentage = 0
        if total_properties - total_booked_properties - total_sold_properties > 0:
            total_occupancy_percentage = round(
                (total_on_lease_properties * 100) / (
                        total_properties - total_booked_properties - total_sold_properties),
                2)
        for col_idx, val in enumerate(['Total', total_properties, total_sold_properties, total_booked_properties, total_on_lease_properties, total_available_properties, f"{total_occupancy_percentage} %"], 1):
            cell = sheet1.cell(row=sheet1_row, column=col_idx, value=val)
            cell.font = bold_font
            cell.fill = orange_fill
            cell.alignment = left_align if col_idx == 1 else right_align
        # Excel Sheet2
        for col_idx, val in enumerate(['Total', f"$ {round(total_ytd_expected, 2)}", f"$ {round(total_ytd_collected, 2)}", f"$ {round(total_ytd_outstanding, 2)}", f"$ {round(total_prev_expected, 2)}", f"$ {round(total_prev_collected, 2)}", f"$ {round(total_prev_outstanding, 2)}"], 1):
            cell = sheet2.cell(row=sheet2_row, column=col_idx, value=val)
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = left_align if col_idx == 1 else right_align

        for rec in sub_properties:
            first_day = today - datetime.timedelta(days=1)
            thirtieth_day = first_day - datetime.timedelta(days=29)
            thirty_first_day = thirtieth_day - datetime.timedelta(days=1)
            sixtieth_day = thirty_first_day - datetime.timedelta(days=29)
            sixty_first_day = sixtieth_day - datetime.timedelta(days=1)
            ninetieth_day = sixty_first_day - datetime.timedelta(days=29)
            ninty_first_day = ninetieth_day - datetime.timedelta(days=1)
            one_twentieth_day = ninty_first_day - datetime.timedelta(days=29)
            invoice_records = invoice.sudo().search(
                [('property_id', '=', rec.id)])
            for customer in invoice_records.mapped('customer_id'):
                one_to_thirty = 0
                thirty_one_to_sixty = 0
                sixty_one_to_ninty = 0
                ninty_one_to_one_twenty = 0
                greater_than_one_twenty = 0
                customer_invoices = invoice.sudo().search(
                    [('customer_id', '=', customer.id),
                     ('property_id', '=', rec.id)])
                for inv in customer_invoices:
                    amount = inv.rent_invoice_id.amount_total
                    if not inv.currency_id.name == 'USD':
                        current_currency = self.env[
                            'res.currency'].sudo().search(
                            [('name', '=',
                              inv.rent_invoice_id.currency_id.name)])
                        us_currency = self.env['res.currency'].sudo().search(
                            [('name', '=', 'USD')])
                        convert_currency = current_currency._convert(amount,
                                                                     us_currency,
                                                                     self.env.company,
                                                                     inv.invoice_date)
                        amount = convert_currency
                    if inv.tenancy_id.contract_type_name == 'lease' and inv.payment_state == 'not_paid':
                        if first_day >= inv.invoice_date >= thirtieth_day:
                            one_to_thirty += amount
                        if thirty_first_day >= inv.invoice_date >= sixtieth_day:
                            thirty_one_to_sixty += amount
                        if sixty_first_day >= inv.invoice_date >= ninetieth_day:
                            sixty_one_to_ninty += amount
                        if ninty_first_day >= inv.invoice_date >= one_twentieth_day:
                            ninty_one_to_one_twenty += amount
                        if one_twentieth_day > inv.invoice_date:
                            greater_than_one_twenty += amount
                if one_to_thirty != 0 or thirty_one_to_sixty != 0 or sixty_one_to_ninty != 0 or ninty_one_to_one_twenty != 0 or greater_than_one_twenty != 0:
                    # Excel Sheet3
                    for col_idx, val in enumerate([customer.name, rec.name, f"$ {round(one_to_thirty, 2)}", f"$ {round(thirty_one_to_sixty, 2)}", f"$ {round(sixty_one_to_ninty, 2)}", f"$ {round(ninty_one_to_one_twenty, 2)}", f"$ {round(greater_than_one_twenty, 2)}"], 1):
                        cell = sheet3.cell(row=sheet3_row, column=col_idx, value=val)
                        cell.alignment = left_align if col_idx <= 2 else right_align
                    sheet3_row += 1
                    a_data = {
                        'customer': customer.name,
                        'property': rec.name,
                        'one_to_thirty': round(one_to_thirty, 2),
                        'thirty_one_to_sixty': round(thirty_one_to_sixty, 2),
                        'sixty_one_to_ninty': round(sixty_one_to_ninty, 2),
                        'ninty_one_to_one_twenty': round(
                            ninty_one_to_one_twenty, 2),
                        'greater_than_one_twenty': round(
                            greater_than_one_twenty, 2)
                    }
                    ageing_data.append(a_data)

        ctx = {
            'properties': property_data,
            'collection_records': collection_data,
            'total_prev_outstanding': round(total_prev_outstanding, 2),
            'total_prev_expected': round(total_prev_expected, 2),
            'total_prev_collected': round(total_prev_collected, 2),
            'total_ytd_collected': round(total_ytd_collected, 2),
            'total_ytd_expected': round(total_ytd_expected, 2),
            'total_ytd_outstanding': round(total_ytd_outstanding, 2),
            'ageing_data': ageing_data
        }
        # For Excel
        stream = BytesIO()
        workbook.save(stream)
        out = base64.encodebytes(stream.getvalue())

        attachment = self.env['ir.attachment'].sudo()
        filename = 'Property Reports' + ".xlsx"
        attachment_id = attachment.sudo().create(
            {'name': filename,
             'type': 'binary',
             'public': False,
             'datas': out})
        # Mail
        mail_values = {
            'email_to': self.env['ir.config_parameter'].sudo().get_param(
                'rental_management.property_report_mails'),
            'email_from': self.env['ir.config_parameter'].sudo().get_param('rental_management.alert_mail_sender_email') or self.env.company.email or '',
            'attachment_ids': [attachment_id.id, ]
        }
        if mail_template:
            mail_template.with_context(ctx).send_mail(self.id, force_send=True,
                                                      email_values=mail_values)

    def _compute_properties(self):
        for rec in self:
            rec.property_count = self.env['property.details'].search_count(
                [('parent_property_id', '=', rec.id),
                 ('is_parent_property', '=', True)])

    def action_properties_parent(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Properties',
            'res_model': 'property.details',
            'domain': [('parent_property_id', '=', self.id),
                       ('is_parent_property', '=', True)],
            'context': {'default_parent_property_id': self.id,
                        'default_is_parent_property': True},
            'view_mode': 'kanban,list,form',
            'target': 'current'
        }


class MeterTypeLines(models.Model):
    _name = "meter.type.lines"
    _description = "For Utility Rate"
    _rec_name = "meter_type_id"

    meter_type_id = fields.Many2one("meter.type", string="Meter Type")
    rate = fields.Monetary(string="Rate")
    currency_id = fields.Many2one('res.currency', string="Currency")
    main_property_id = fields.Many2one('parent.property',
                                       string="Main Property")


class FloorPlan(models.Model):
    _name = 'floor.plan'
    _description = 'Details About Floor Plan'

    image = fields.Image(string='Image')
    title = fields.Char(string='Title')
    property_id = fields.Many2one('property.details', string='Property')


class PropertyImages(models.Model):
    _name = 'property.images'
    _description = 'Property Images'

    property_id = fields.Many2one(
        'property.details', string='Property Name', readonly=True)
    title = fields.Char(string='Title')
    image = fields.Image(string='Images')
    parent_property_id = fields.Many2one('parent.property',
                                         string="Parent Property")


class PropertyTag(models.Model):
    _name = 'property.tag'
    _description = 'Property Tags'
    _rec_name = 'title'

    title = fields.Char(string='Title')
    color = fields.Integer(string='Color')


class TenancyExtraService(models.Model):
    _inherit = 'product.product'

    is_extra_service_product = fields.Boolean(string="Is Extras Service")


class ExtraServiceLine(models.Model):
    _name = 'extra.service.line'
    _description = "Tenancy Extras Service"

    service_id = fields.Many2one('product.product', string="Service", domain=[
        ('is_extra_service_product', '=', True)])
    price = fields.Float(related="service_id.lst_price", string="Cost")
    service_type = fields.Selection(
        [('once', 'Once'), ('monthly', 'Monthly')], string="Type",
        default="once")
    property_id = fields.Many2one('property.details', string="Property")


class PropertyResCity(models.Model):
    _name = 'property.res.city'
    _description = 'Cities'
    name = fields.Char(string="City Name", required=True)


class PropertyConnectivity(models.Model):
    _name = 'property.connectivity'
    _description = "Property Nearby Connectivity"

    name = fields.Char(string="Title")
    distance = fields.Char(string="Distance")
    image = fields.Image(string='Images')


class PropertyConnectivityLine(models.Model):
    _name = 'property.connectivity.line'
    _description = "Property Connectivity Line"

    property_id = fields.Many2one('parent.property')
    connectivity_id = fields.Many2one(
        'property.connectivity', string="Nearby Connectivity")
    image = fields.Image(related="connectivity_id.image", string='Images')
    distance = fields.Char(string="Distance")


class TenancyInquiry(models.Model):
    _name = 'tenancy.inquiry'
    _description = "Tenancy Inquiry"
    _rec_name = 'lead_id'

    property_id = fields.Many2one(
        'property.details', string="Property Details")
    note = fields.Text(string="Note")
    duration_id = fields.Many2one('contract.duration', string='Duration')
    customer_id = fields.Many2one('res.partner', string="Customer")
    lead_id = fields.Many2one('crm.lead', string="Lead")

    def _compute_display_name(self):
        for rec in self:
            if rec.lead_id:
                rec.display_name = '%s - %s' % (rec.customer_id.name, rec.lead_id.name)
            else:
                rec.display_name = rec.customer_id.name or ''


class SaleInquiry(models.Model):
    _name = 'sale.inquiry'
    _description = "Sale Inquiry"
    _rec_name = 'lead_id'

    property_id = fields.Many2one(
        'property.details', string="Property Details")
    note = fields.Text(string="Note")
    company_id = fields.Many2one(
        'res.company', string='Company', default=lambda self: self.env.company)
    currency_id = fields.Many2one(
        'res.currency', related='company_id.currency_id', string='Currency')
    ask_price = fields.Monetary(string="Ask Price")
    customer_id = fields.Many2one('res.partner', string="Customer")
    lead_id = fields.Many2one('crm.lead', string="Lead")

    def _compute_display_name(self):
        for rec in self:
            if rec.lead_id:
                rec.display_name = '%s - %s' % (rec.customer_id.name, rec.lead_id.name)
            else:
                rec.display_name = rec.customer_id.name or ''


class PropertyAssets(models.Model):
    _name = 'property.assets'
    _description = 'Property Assets'

    handover_property_id = fields.Many2one('handover.property',
                                           string="Handover Property")
    property_id = fields.Many2one('property.details', string="Property")
    display_type = fields.Selection(
        selection=[('line_section', "Section"), ('line_note', "Note")],
        default=False)
    name = fields.Text(string="Description")
    serial_no = fields.Integer(string="Sr.No", compute="_compute_sr_no")
    product_id = fields.Many2one('product.product', string="Product")
    Prod_qty = fields.Float(string="Qty", default="1.0")
    prod_condition = fields.Selection([('new', 'New'), ('used', 'Used')],
                                      string="Condition")
    remark = fields.Text(string="Remark")

    def _compute_sr_no(self):
        no = 0
        for line in self:
            if line.display_type not in ['line_section']:
                no += 1
                line.serial_no = no
            else:
                line.serial_no = 0
                no = 0

    @api.onchange('product_id')
    def _onchange_prod_desc(self):
        for rec in self:
            if rec.product_id:
                rec.name = rec.product_id.get_product_multiline_description_sale()
