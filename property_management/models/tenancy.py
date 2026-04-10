# -*- coding: utf-8 -*-
# Copyright 2020-Today TechKhedut.
# Part of TechKhedut. See LICENSE file for full copyright and licensing details.
import base64
import re
from datetime import timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

from odoo import api, fields, models, _
from odoo.exceptions import UserError


class ContractSubStage(models.Model):
    _name = 'contract.sub.stage'
    _description = 'Contract Sub Stage'

    name = fields.Char(string='Name', required=True)


class TenancyDetails(models.Model):
    _name = 'tenancy.details'
    _description = 'Information Related To customer Tenancy while Creating Contract'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _rec_name = 'tenancy_seq'

    def _get_default_sub_stage_id(self):
        return self.env['contract.sub.stage'].search([], limit=1)

    # @api.depends('tenancy_id', 'contract_type_name', 'property_id', 'total_rent', 'rent_unit', 'start_date', 'end_date',
    #              'security_deposite')
    # def _get_content(self):
    #     for rec in self:
    #
    #         if rec.property_id:
    #             cur_full_name = rec.property_id.currency_id.full_name
    #             amt_in_text = rec.property_id.currency_id.with_context(lang='en').amount_to_text(
    #                 rec.property_id.tenancy_price)
    #             cur_name = rec.property_id.currency_id.name
    #             tenancy_price = rec.property_id.tenancy_price
    #             rent_unit = rec.property_id.rent_unit
    #         else:
    #             cur_full_name = ''
    #             amt_in_text = ''
    #             cur_name = ''
    #             tenancy_price = 0.0
    #             rent_unit = ''
    #
    #         if rec.start_date:
    #             start_date = rec.start_date.strftime('%d/%B/%Y')
    #         else:
    #             start_date = False
    #
    #         if rec.end_date:
    #             end_date = rec.end_date.strftime('%d/%B/%Y')
    #         else:
    #             end_date = False
    #
    #         vals = ''
    #         vals += '''
    #             <!DOCTYPE html>
    #             <html>
    #             <head>
    #                 <title>Sample Report</title>
    #             </head>'''
    #         vals += '''<body> '''
    #         vals += '''
    #                 <div class="page" style="page-break-before: always;">
    #                     <h2 style="text-align:center">
    #                         AGREEMENT FOR LEASE
    #                     </h2>
    #
    #                     <h4 style="text-align:center;padding-top:110px;">
    #                         Between
    #                     </h4>
    #
    #
    #                     <h2 style="text-align:center;padding-top:110px;">
    #                         <span>%s</span>
    #                     </h2>
    #
    #                     <h5 style="text-align:center;"><span>‘The Landlord’</span></h5>
    #
    #                     <h4 style="text-align:center;padding-top:110px;">
    #                         And
    #                     </h4>
    #
    #                     <h2 style="text-align:center;padding-top:110px;">
    #                         <span>%s</span>
    #                     </h2>
    #
    #                     <h5 style="text-align:center;"><span>‘The Tenant’</span></h5>
    #
    #                     <div name="company_address" style="padding-top:130px;font-size:18px;">
    #                         <div>
    #                             <span style="text-decoration:underline;font-weight:bold;">Drawn By:</span>
    #                         </div>
    #                         <div>
    #                             %s
    #                         </div>
    #                         <div t-if="rec.property_id.pro_company_id.zip">P.O Box %s,</div>
    #                         <div t-if="rec.property_id.pro_company_id.street">%s.</div>
    #                     </div>
    #
    #                     <br/><br/><br/><br/><br/><br/>
    #                 </div>''' % (rec.property_landlord_id.name or '', rec.tenancy_id.name or '',
    #                              rec.property_id.pro_company_id.name or '',
    #                              rec.property_id.pro_company_id.zip or '',
    #                              rec.property_id.pro_company_id.street or ''
    #                              )
    #
    #         vals += '''<div style="font-size:18px;">'''
    #         vals += '''
    #                 <div>
    #                     <h4 style="text-align:center;text-decoration:underline;">AGREEMENT FOR LEASE OF APARTMENT</h4>
    #                 </div>
    #
    #                 <div style="padding-top:10px;">
    #                     <strong>THIS AGREEMENT</strong> is made at Dar es Salaam this …………………………….. day of ……………………2023
    #                 </div>
    #
    #                 <div style="text-align:center;padding-top:15px;font-weight:bold;">
    #                     <h5>BETWEEN</h5>
    #                 </div>
    #
    #                 <div style="padding-top:10px;">
    #                     <p><strong>M/S %s </strong>
    #                     a private limited liability company registered in Tanzania whose registered office is at <b>Plot No</b>.40 Mikocheni Light Industrial Area, Dar es Salaam, and of Post office Box 19716, Dar es salaam, Tanzania. (Hereinafter referred to as ‘the <b>Landlord,</b>’ which expression shall include unless inconsistent with the context the successors and assigns  in title)  of the one part.
    #                     </p>
    #                 </div>
    #
    #                 <div style="text-align:center;padding-top:15px;font-weight:bold;">
    #                     <h5>AND</h5>
    #                 </div>
    #
    #                 <div style="padding-top:10px;">
    #                     <p>
    #                         <strong>%s</strong>,of <b>P.O Box No:%s %s, %s</b>.(Hereinafter referred to as ‘the <b>  Tenant,</b>’ which expression shall include unless inconsistent with the context the successors and assigns in title) of the other part.
    #                     </p>
    #                 </div>
    #
    #                 <div style="padding-top:15px;font-weight:bold;">
    #                     <h5>WHEREAS</h5>
    #                 </div>
    #
    #                 <div style="margin-left:25px;">
    #                     <p>
    #                         The <b>Landlord</b> being a developer is the owner in occupation of a building with Flat Apartments on Plot No. <b>%s %s</b> together with all the unexhausted improvements, developments and appurtenances therein contained (herein after referred to as ‘the <b>Property</b>’).
    #                     </p>
    #
    #                     <p>
    #                         The Landlord has offered to lease a flat apartment No <b>%s</b> of Tower <b>%s</b> on the <b>%s</b> floor of the Property (herein after referred to as the ‘<b>Demised Premises</b>’) and the Tenant has agreed to enter into a lease agreement of the said Demised Premises subject to the herein below undertaking.
    #                     </p>
    #                 </div>''' % (
    #             rec.property_landlord_id.name or '', rec.tenancy_id.name or '', rec.tenancy_id.zip or '',
    #             rec.tenancy_id.street or '',
    #             rec.tenancy_id.country_id.name or '', rec.property_id.plot_no or '', rec.property_id.street or '',
    #             rec.property_id.room_no or '', rec.property_id.no_of_towers or '', rec.property_id.floor or '')
    #
    #         vals += '''
    #
    #                 <div style="padding-top:10px;">
    #                     <p>Now this Agreement Witnesseth As follows:</p>
    #                     <p> Operative Provisions:</p>
    #                     <p> 1. Definations and Interpretation</p>
    #                     <div style="margin-left:20px;">
    #                         <p>(a)Definitions and Interpretation</p>
    #                         <div style="margin-left:20px;">
    #                             <p>(i)In this Agreement if the Context so allows:</p>
    #                             <p>‘<b>Agreement</b>’ means this Lease Agreement, rules and regulations, annextures and includes any amendment or other novation agreed and duly signed and executed by the parties in accordance with the terms of this Agreement.</p>
    #                             <p>‘<b>Co-Owners</b>’ means the other occupiers either by rent or purchase of the flat apartments in the Property.</p>
    #
    #                             <p>‘<b>Common area</b>’ means whole or a portion of property that is used by all co- owners for their common use.</p>
    #
    #                             <p>‘<b>Demised Premise</b>’ means the flat apartment offered to the Tenant as it is and basis and includes the parking lot allocated.</p>
    #
    #                             <p>‘<b>Encumbrance</b>’ means any mortgage, charge, pledge, lien, assignment, hypothecation, preferential right or trust, arrangement or other encumbrance, security, agreement or arrangement of any kind or any right conferring a priority of payment affecting the title to the property.</p>
    #
    #                             <p>‘<b>Management</b>’ means the estate manager incharge of the day to day management of the common property.</p>
    #
    #                             <p>‘<b>Monthly service charges</b>’ means the charges made on items listed on annexure A1 only.</p>
    #
    #                             <p>‘<b>Property</b>’ means Part of all the piece of land known as Plot no’s
    #                             <b>%s %s</b>, together with the storeway building where the Demised Premises is part of the flat apartment.</p>
    #
    #                             <p>‘<b>Security Deposit</b>’ means an amount deposited by the Tenant to the Landlord as security for any damages or defaults, its refundable at the termination of the Lease Agreement in accordance with the Tenants covenants and obligations.</p>
    #
    #                             <p>‘<b>Vendor</b>’ means the seller of the Demised premise and Landlord.</p>
    #                         </div>
    #
    #                         <p>
    #                             (b) Words of one gender include both other genders, and entities, and words denoting natural persons include corporations and firms.
    #                         </p>
    #                         <p>
    #                             (c) Words denoting an obligation on a party to do any acts, matter or thing include an obligation to procure that it be done, and words placing a party under a restriction include an obligation not to permit infringement of the restriction.
    #                         </p>
    #                         <p>
    #                             (d) Where the Tenant comprises two or more parties, the obligations of the Tenant are in relation to each such party joint and several.
    #                         </p>
    #                         <p>
    #                             (e) Reference to ‘liability’ include where the context so allows, claims, demands, proceedings, damages, costs and expenses.
    #                         </p>
    #                     </div>
    #                 </div>''' % ((rec.property_id.plot_no or ''), (rec.property_id.street or ''))
    #
    #         vals += '''<div style="padding-top:10px;"> '''
    #         vals += '''<p>NOW THIS Agreement Witnesseth as follows:</p>'''
    #         vals += '''<div style="margin-left:20px;">'''
    #         vals += '''
    #                 <p>1. Lease</p>
    #                 <div style="margin-left:20px;">
    #                     <p>
    #                         In pursuance of the said Agreement and in Consideration of the rent hereby reserved and the Tenant’s Covenants hereinafter mentioned, the Landlord hereby demises unto the Tenant the Demised Premises for a period of <b>%s</b> Commencing on The  <b>%s</b>, and ending on  <b>%s</b>, subject to renew on terms and conditions as hereunder.
    #                     </p>
    #                 </div>''' % (rec.duration_id.duration or '', start_date, end_date)
    #
    #         vals += '''
    #                 <p>2. Payment of the Rent.</p>
    #                 <div style="margin-left:20px;">
    #                     <p><b>(i)</b>  The rent herein reserved shall be  <b>%s. %s. (%s %s /- only)  %s</b>, to be paid three months in advance amounting to total amount of USD. Four Thousand Nine Hundred Fifty Only. (USD 4,950/- Only) + VAT (If Applicable). The said rental amount must be paid timely as agreed and any late payment shall attract a monthly interest rate of <b>%s %% </b>
    #                     of monthly rent on each month of delay.
    #                     </p>
    #
    #                     <p><b>(ii)</b>  The Tenant shall be responsible for payment of stamp duty on this Lease Agreement and its Counterpart,
    #                     </p>
    #
    #                     <p><b>(iii)</b>  Tenant shall deduct the Withholding Tax at the prevailing rate (Presently <b>%s %% </b>) as required under section 82 of Income Tax Act and must produce to the Landlord the Withholding Certificate in the manner prescribed under Sec. 85 of Income Tax Act with proof of payment.
    #                     </p>
    #
    #                     <p>
    #                        For claiming this withholding tax, Landlord has every right to ensure that Tenant has made the payment of Withholding Tax to the Tanzania Revenue Authority.
    #                     </p>
    #
    #                     <p>
    #                         For avoidance of doubt, the Tenant must provide the Withholding Tax Certificate and proof of payment within thirty days from the end of the month in which deduction is made, failure of which the landlord shall exercise his powers and recover the same amount from the Tenant
    #                     </p>
    #
    #                     <p><b>(iv)</b>  Value Added Tax (VAT), is to be added. </p>
    #
    #                     <p><b>(v)</b>  To pay the monthly service charge as per the Property management contract entered by the parties and the estate management as part of this agreement. </p>
    #
    #                     <p><b>(vi)</b>  The service charges and other costs/reimbursable must be paid timely and any default shall amount to penalties not limited to termination of the lease agreement upon issuing a default notice to the Tenant.
    #                     </p>
    #
    #                     <p><b>(vii)</b>  Any late payments of service charges and utility charges shall amount to disconnection/discontinue of the Services offered by the Property manager upon issuing a seventy-two (72) hours default notice.
    #                     </p>
    #                 </div>''' % (cur_full_name, amt_in_text,
    #                              cur_name, tenancy_price,
    #                              rent_unit, rec.monthly_interest_rate,
    #                              rec.presently_rate)
    #
    #         vals += '''
    #                 <p>3. Landlord’s Warranties and Representations.</p>
    #                 <div style="margin-left:20px;">
    #                     <p>
    #                         The Landlord warrants and represents to the Tenant that it is the legal owner in possession of the property herein and that it is legally capable and duly authorized to enter into this lease agreement and perform all the obligations set out herein.
    #                     </p>
    #                 </div>'''
    #
    #         vals += '''
    #                 <p>4. Tenant’s Rights and Obligations.</p>
    #                 <div style="margin-left:20px;">
    #                     <p>
    #                         The Tenant hereby covenant with the Landlord as follows:
    #                         <b>4.1</b> To pay the monthly Rent in advance as agreed in clause 2 of this agreement.
    #                     </p>
    #
    #                     <p><b>4.2</b> To pay the service charges in advance and on timely basis.</p>
    #
    #                     <p><b>4.3</b> To pay all fees and all rates and charges for use of water from Dawasco or any other body that may be in force, electricity from Tanesco or any other power generating body that may be in force in respect of demised premises during the said lease period.</p>
    #
    #                     <p><b>4.3.1</b> Failure to comply with the payments of such services shall amount to disconnection of the services for the entire period and any late payments shall be made with interests and penalties.</p>
    #
    #                     <p><b>4.4</b> At all times to keep the interior of the demised premises and the appurtenances thereof including doors, windows and other fixtures, fittings, fastening, electrical wires, water drains, water fittings and other pipes and paintings and decorations thereof in  substantial routine repair and good condition.
    #                     </p>
    #
    #                     <p><b>4.5</b> To keep the surrounding ground of the demised premises in good order and condition.
    #                     </p>
    #
    #                     <p><b>4.6</b> To use the demised premises for residential purposes only and maintain a high moral ethical standards.</p>
    #
    #                     <p><b>4.7</b> To permit the Landlord and or his agents with or without workmen at all reasonable times during day time by prior appointment to enter upon the demised premises for the purpose of viewing and executing any repair necessary under the covenants herein contained.
    #                     </p>
    #
    #                     <p><b>4.8</b> Not to make any alterations or additions to the demised premises without first obtaining the written consent of the Landlord.</p>
    #
    #                     <p><b>4.9</b> Not to assign or sublet or part with the possession of the demised premises or any part thereof without the written consent of the Landlord, PROVIDED ALWAYS that the occupation of the demised premises or any part thereof by any person in the service or employment of the Tenant shall not constitute an assignment, underlies or parting with the possession of the demised premises or any part thereof.</p>
    #
    #                     <p>
    #                         Incase  the  Tenant  decides  to  sub-lease,  the  terms  and  conditions  of  this  lease agreement together with the rules and regulations must be followed.
    #                     </p>
    #
    #                     <p><b>4.10</b> Not to use the demised premises in a way that would create annoyance, disturbance, public/private threat, nuisance or any danger to the public, neighbours or other Tenants.</p>
    #
    #                     <p><b>4.11</b> Persons of indecent character such as prostitutes, drug dealers or use of the Demised premise as a brothel is strictly not allowed in the Demised premises such an act shall amount to immediate termination of this Agreement.</p>
    #
    #                     <p><b>4.12</b> To insure and keep insured himself, family, employees, his invitees and all his/ their personal belongings both in the demised premises and the Common  Area against loss or damage by fire and any other risks.</p>
    #
    #                     <p><b>4.13</b> To indemnify the Landlord for any loss that he may sustain as a result of failure by the Tenant to maintain adequate insurance for its property.</p>
    #
    #                     <p><b>4.14</b> The Landlord shall not be responsible for any damage, theft or loss whatsoever involving the Tenant’s properties/belongings as the security shall be the responsibility of the Tenant.</p>
    #
    #                     <p><b>4.15</b> The Tenant shall provide adequate security for the property and all goods kept on the demised premise including his parking place, Tenant agrees that the Landlord SHALL not be liable in any way whatsoever, for any theft or loss on the demised premise during the entire lease period.</p>
    #
    #                     <p><b>4.16</b> That the Tenants hereby agrees that any domestic workers/house help whose engaged to offer services to the tenants, and who resides outside the premises shall be subjected to the inspection by the security guard during the reporting time and departing/leaving time at the entrance/exit gate.</p>
    #
    #                     <p><b>4.17</b> On the expiration or sooner after termination of the term hereby granted, to deliver up the demised premises to the Landlord in good tenable condition and leave all the fixtures and furniture’s of the property intact and having been repaired reasonable wear and tear excepted.</p>
    #
    #                     <p><b>4.18</b> Tenant shall also  pay  to  Landlord  a  “<b>Security  Deposit</b>”  of  <b>%s%s</b>/- as security for any damages or defaults.</p>
    #
    #                     <p><b>4.19</b> Should any damage, breakage or non-repair on the demised premises occur and the Tenant fails to make good of the same, the Landlord shall assess the costs of such damages and thereby notify the Tenant to make good of the same, failure of which the Landlord shall use the security deposit to repair or replace the items.</p>
    #
    #                     <p><b>4.20</b> To pay for the fuel consumption of the generator used when electricity from the national grid is interrupted on a pro-rata basis or basing on metric basis and upon issuance of a reasonable notice from the Landlord, failure of which amounts to the breach of this agreement and disconnection of the services.</p>
    #
    #                     <p><b>4.21</b> That the Tenant, his agents, associates, sub lessees and any other occupiers of the demised premises shall abide to the rules and regulations of the demised premises and apartments failure of which shall led to the breach of this contract and the Landlord’s entitlement to remedies.</p>
    #
    #                     <p><b>4.22</b> That the Tenant has to read, understand and observe the rules and regulations herein attached to be read together with this agreement.</p>
    #                 </div>''' % (rec.currency_id.symbol or '', rec.security_deposite or 0.0)
    #
    #         vals += '''
    #                 <p>5. Landlord’s rights and obligations.</p>
    #                 <div style="margin-left:20px;">
    #                     <p><b>5.1</b> During the subsistence of this agreement not to sell, assign, transfer, lease, sublet or otherwise dispose and deal with the demised premises in a manner prejudicial to the Tenant’s rights contained under this agreement.</p>
    #
    #                     <p><b>5.2</b> During the subsistence of the lease, the Landlord will be responsible for payment of property tax and all the site rates, land rent or other statutory charges and impositions levied in respect of the Demised Premise to the property owners by the government during the currency of the said term.</p>
    #
    #                     <p><b>5.3</b> To permit the Tenant hereby observing and performing the agreed covenants to peacefully posses the demised premises during the lease period.</p>
    #
    #                     <p><b>5.4</b> To provide the Tenant a parking place for only one car which shall only be used for parking the Tenant’s car only, the allocated parking shall not be used for any other purposes other than parking the Tenant’s car.</p>
    #
    #                     <div style="margin-left:20px;">
    #                         <p><b>5.5.1</b> The parking is restricted to the labelled one and not otherwise.</p>
    #
    #                         <p><b>5.5.2</b> Any obstruction or wrong parking shall amount to trespass with consequences as stayed in the rules and regulation.</p>
    #
    #                         <p><b>5.5.3</b> Parking allocated shall be at car owner’s risk and not the liability of the Landlord or the management.</p>
    #
    #                         <p><b>5.5.4</b> No services or maintenance of the Tenant’s vehicle in the property and in the parking area unless it is an emergency cases which must be reported to the Property manager.</p>
    #                     </div>
    #
    #                     <p><b>5.5</b> If at any time the demised premises are rendered unfit for use by accidental fire or other causes beyond the control of the Tenant, the Landlord shall allow the Tenant an abatement of all or part of the said rent proportionate to the existence and duration of the damage until the demised premises shall be fit for use again.</p>
    #                 </div>'''
    #
    #         vals += '''
    #                 <p>6. PROVIDED ALWAYS and it is hereby agreed and declared that.</p>
    #                 <div style="margin-left:20px;">
    #                     <p><b>6.1</b> In the event the Tenant fails to fulfil any of its obligations under this lease, and where this lease specifically provides no other remedy for such failure, the Landlord shall give the Tenant notice in writing to remedy the breach within a period of not less than fourteen (14) days. If upon such notice, the Tenant fails to remedy the default, the Landlord will be entitled to issue thirty (30) days notice of termination of the lease.</p>
    #
    #                     <p>Upon such termination, the Landlord shall refund to the Tenant any Rent received in advance for an unexpired period of the Lease, without interests and after deducting one month’s rent as notice of default and all taxes paid and costs for any damages assessed by the Parties.</p>
    #
    #                     <p><b>6.2</b> If and whenever during the said lease period the said rent or any part thereof shall be unpaid for 30 days after the date due for payment and if any covenant on the Tenants part herein contained shall not be performed or observed then in any of the said cases it shall be lawful for the Landlord at any time thereafter to re-enter upon the property or any part thereof and then  this  lease  agreement  shall  absolutely  determine  without  prejudice to  any  rights  or remedies which may have accrued to either party against the other in respect of antecedent breach of any of the covenants herein contained and the Landlord shall have the right to take any legal action in recovery of any rent due and accruing.</p>
    #
    #                     <p><b>6.3</b>The Tenant and the Landlord shall renew this Lease Agreement by giving each other a written notice of thirty days before the expiration of lease period in terms granted herein expressing their intention to agree to the terms and conditions of the renewed lease period. If the parties agree, the renewed lease period shall commence immediately upon the expiration of the current lease period.</p>
    #
    #                     <p><b>6.4</b> (i)Either Party may by giving thirty (30) days written notice assigning reasons, terminate this agreement, reasons which should not be unreasonably withheld and any advance amount paid for the unexpired period of the Lease shall be refunded after deducting all the taxes paid and any damages assessed.</p>
    #
    #                     <p>Any termination by the Tenant before the expiry of the Lease period without giving thirty days’ notice shall amount to deduction by the Landlord of one month’s rent from the security deposit or one month’s rent deduction from any rent paid for the unexpired period in lieu of the notice.</p>
    #
    #                     <p><b>6.4</b> (ii)Incase of the Tenant’s premature termination of the  lease  agreement  before  the expiry of the first Six months or before quarter of the period of the Lease Agreement, advance rent paid for first six months or quarter of the period of the Lease Agreement shall not be refundable.</p>
    #
    #                     <p><b>6.5</b> (i)At the time of termination or at the expiry (maturity) of the lease agreement, The keys and all other properties of the landlord MUST be handed over to the duly authorized property Manager.</p>
    #
    #                     <p><b>6.5</b> (ii)Any delay in the handing over of the key shall amount to penalty which shall attract an amount equivalent to one month’s lease.</p>
    #
    #                     <p><b>6.5</b> (iii)Upon handing over of the keys or within thirty days before the expiry or termination of the contract a formal inspection of the demised premises shall be done in presence of both the tenant and property manager, Any damages shall be assessed and its only after a formal clearance by the property management that the Tenant shall be allowed to move his/her properties.</p>
    #
    #                     <p><b>6.6</b> Any notice under this lease agreement shall be in writing and any notice to the Tenant shall be sufficiently served if left addressed to him on the demised premises or served to him by registered post and any notice to the Landlord shall be sufficiently served if sent to him by registered post addressed to it at its usual place of business.</p>
    #
    #                     <p><b>6.7</b> Recreation areas and facilities on the third floor of the said building shall be reserved for the members only upon registration.</p>
    #                 </div>'''
    #
    #         vals += '''
    #                 <p>7. Dispute Resolution.</p>
    #                 <div style="margin-left:20px;">
    #                     <p>This Agreement shall be construed and governed by the Laws of Tanzania and any dispute, difference arising out of or in connection with this Agreement shall be amicably mediated between the parties within thirty days from the occurrence of the dispute. Failure of which the matter shall be referred to the court of competent jurisdiction with in the united republic of Tanzania.</p>
    #                 </div>'''
    #
    #         vals += '''
    #                 <p><b>IN WITNESS WHEREOF</b> the parties hereto have duly executed these presents in the manner and on the day and year hereinafter appearing.</p>'''
    #
    #         vals += '''
    #                 <p>Dated at Dar es Salaam this .............day of ............... 2023</p>
    #                 '''
    #
    #         vals += '''
    #
    #                 <div style="padding-top:10px;">
    #                     <div class="row">
    #                         <div class="col-6 text-left">
    #                             <p>
    #                                 <b>SEALED</b> with the <b>Common Seal</b> of the said
    #                                 <b>M/S <span>%s</span></b>. and DELIVERED in presence of us this………..……………day of………………………2023
    #                             </p>
    #                         </div>
    #                         <div class="col-2 text-left"/>
    #                         <div class="col-4 text-right">
    #                                 <div class="text-center">_______________________</div>
    #                                 <div class="text-center" style="font-weight:bold;padding-top:5px;">LANDLORD</div>
    #                         </div>
    #                     </div>
    #
    #                     <div class="row" style="font-weight:bold;padding-top:20px;">
    #                         <div class="col-6">
    #                             <div style="padding-top:5px;">Signature: ………………………………………………
    #                             </div>
    #                             <div style="padding-top:5px;">Full Name: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Address: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Qualification: ………………………………………………</div>
    #                         </div>
    #                     </div>
    #
    #                     <div class="row" style="font-weight:bold;padding-top:20px;">
    #                         <div class="col-6">
    #                             <div style="padding-top:5px;">Signature: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Full Name: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Address: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Qualification: ………………………………………………</div>
    #                         </div>
    #                     </div>
    #                 </div>''' % (rec.property_landlord_id.name or '')
    #
    #         vals += '''
    #                 <div style="padding-top:15px;">
    #                     <div class="row">
    #                         <div class="col-6 text-left" style="padding-top:20px;">
    #                             <p>
    #                                 <b>SEALED</b> with the <b>Common Seal</b> of the said
    #                                 <b>M/S <span>%s</span></b>. and DELIVERED in presence of us this………..……………day of………………………2023
    #                             </p>
    #                         </div>
    #                         <div class="col-2 text-left"/>
    #                         <div class="col-4 text-right">
    #                                 <div class="text-center">_______________________</div>
    #                                 <div class="text-center" style="font-weight:bold;padding-top:5px;">TENANT</div>
    #                         </div>
    #                     </div>
    #
    #                     <div class="row" style="font-weight:bold;padding-top:20px;">
    #                         <div class="col-6">
    #                             <div style="padding-top:5px;">Signature: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Full Name: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Address: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Qualification: ………………………………………………</div>
    #                         </div>
    #                     </div>
    #
    #                     <div class="row" style="font-weight:bold;padding-top:20px;">
    #                         <div class="col-6">
    #                             <div style="padding-top:5px;">Signature: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Full Name: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Address: ………………………………………………</div>
    #                             <div style="padding-top:5px;">Qualification: ………………………………………………</div>
    #                         </div>
    #                     </div>
    #                 </div>''' % (rec.tenancy_id.name or '')
    #
    #         vals += '''</div></div></div> '''
    #         vals += '''</body>
    #                 </html>
    #
    #             '''
    #
    #         self.contract_terms = vals

    contract_terms = fields.Html(string='Contract Term and Condition')
    agreement_template_id = fields.Many2one('agreement.template',
                                            "Agreement Template")

    tenancy_seq = fields.Char(string='Sequence', required=True, readonly=True,
                              copy=False, default=lambda self: ('New'))
    company_id = fields.Many2one('res.company', string='Company',
                                 default=lambda self: self.env.company,
                                 tracking=True)
    pro_company_id = fields.Many2one('res.company', string='Property Company',
                                     default=lambda self: self.env.company,
                                     tracking=True)
    pro_management_company_id = fields.Many2one('res.company',
                                                string='Property Management Company',
                                                tracking=True)

    currency_id = fields.Many2one('res.currency', string='Currency',
                                  tracking=True)

    close_contract_state = fields.Boolean(string='Contract State')
    active_contract_state = fields.Boolean(string='Active State')
    is_extended = fields.Boolean(string='Extended', tracking=True)
    contract_sub_type_id = fields.Many2one('contract.sub.stage',
                                           string='Contract Sub Type',
                                           default=_get_default_sub_stage_id,
                                           copy=False,
                                           store=True, tracking=True)

    contract_type = fields.Selection([('new_contract', 'Draft'),
                                      ('running_contract', 'Running'),
                                      ('cancel_contract', 'Cancel'),
                                      ('close_contract', 'Terminated'),
                                      ('expire_contract', 'Expired')],
                                     string='Contract Type',
                                     default="new_contract", tracking=True)

    # Tenancy Information
    tenancy_id = fields.Many2one('res.partner', string='Tenant',
                                 domain=[('is_customer', '=', True)],
                                 tracking=True)
    is_any_broker = fields.Boolean(string='Any Broker', tracking=True)
    broker_id = fields.Many2one('res.partner', string='Broker', tracking=True)
    commission = fields.Monetary(string='Commission ',
                                 compute='_compute_broker_commission',
                                 store=True, tracking=True)
    last_invoice_payment_date = fields.Date(string='Last Invoice Payment Date',
                                            tracking=True)
    broker_invoice_state = fields.Boolean(string='Broker  invoice State',
                                          tracking=True)
    broker_invoice_id = fields.Many2one('account.move', string='Bill',
                                        tracking=True)
    term_condition = fields.Html(string='Term and Condition')
    agreement = fields.Html(string="Agreement")
    is_any_deposit = fields.Boolean(string="Deposit", tracking=True)
    deposit_amount = fields.Monetary(string="Security Deposit Amount",
                                     tracking=True)
    type = fields.Selection(
        [('automatic', 'Auto Create Rent Invoice Line'),
         ('manual', 'Manually create list of all rent invoice line')],
        default='automatic')

    # Property Information
    property_id = fields.Many2one('property.details', string='Property',
                                  domain=[('stage', '=', 'available')],
                                  tracking=True)
    is_extra_service = fields.Boolean(related="property_id.is_extra_service",
                                      string="Any Extra Services",
                                      tracking=True)
    property_landlord_id = fields.Many2one(related='property_id.landlord_id',
                                           string='Landlord', store=True,
                                           tracking=True)
    property_type = fields.Selection(related='property_id.type',
                                     string='Property Type', store=True,
                                     readonly=True,
                                     tracking=True)
    total_rent = fields.Monetary(string='Rent', tracking=True)
    rent_smtr = fields.Monetary(string='Rent/Square Meter', tracking=True)
    extra_services_ids = fields.One2many('tenancy.service.line', 'tenancy_id',
                                         string="Services", tracking=True)
    payment_term_id = fields.Many2one('contract.duration',
                                      string='Payment Term Id', tracking=True)

    # Time Period
    payment_term = fields.Selection([('monthly', 'Monthly'),
                                     ('full_payment', 'Full Payment'),
                                     ('quarterly', 'Quarterly'),
                                     ('year', 'Yearly')],
                                    string='Payment Term', tracking=True)
    duration_id = fields.Many2one('contract.duration', string='Duration',
                                  tracking=True)
    contract_agreement = fields.Binary(string='Contract Agreement')
    signed_agreement = fields.Binary(string='Signed Contract')
    file_name = fields.Char(string='File Name', tracking=True)
    signed_file_name = fields.Char(string='Signed File Name', tracking=True)
    month = fields.Integer(related='duration_id.month', string='Month',
                           tracking=True)
    start_date = fields.Date(string='Start Date', default=fields.Date.today(),
                             tracking=True)
    end_date = fields.Date(string='End Date', tracking=True)

    rent_type = fields.Selection(
        [('once', 'One Month'), ('e_rent', 'All Month')],
        string='Brokerage Type',
        tracking=True)
    commission_type = fields.Selection([('f', 'Fix'), ('p', 'Percentage')],
                                       string="Commission Type", tracking=True)
    broker_commission = fields.Monetary(string='Broker Commission',
                                        tracking=True)
    broker_commission_percentage = fields.Float(string='Percentage',
                                                tracking=True)

    # Related Field
    rent_invoice_ids = fields.One2many('rent.invoice', 'tenancy_id',
                                       string='Invoices', tracking=True)

    # Tenancy Calculation
    total_tax = fields.Monetary(string="Total Tax",
                                compute="_compute_tenancy_calculation",
                                tracking=True)

    total_tenancy = fields.Monetary(string="Total Tenancy",
                                    compute="_compute_tenancy_calculation",
                                    tracking=True)
    paid_tenancy = fields.Monetary(string="Paid Tenancy",
                                   compute="_compute_tenancy_calculation",
                                   tracking=True)
    remain_tenancy = fields.Monetary(string="Remaining Tenancy",
                                     compute="_compute_tenancy_calculation",
                                     tracking=True)
    rent_unit = fields.Selection(related="property_id.rent_unit",
                                 readonly=False, tracking=True)

    # new fields
    contract_type_name = fields.Selection([('lease', 'Lease Agreement'),
                                           ('service',
                                            'Property Management Agreement')],
                                          string='Contract Type Name',
                                          tracking=True)
    security_deposite = fields.Monetary(string='Security Deposit',
                                        tracking=True)
    withholding_tax = fields.Float(string="Withholding Tax %", tracking=True)
    withholding_tax_paid = fields.Many2one('res.partner',
                                           string="Withholding Tax Paid by",
                                           tracking=True,
                                           domain=[('is_customer', '=', True)])
    days_left = fields.Integer(string="Days Left", compute="_get_days",
                               default=0, tracking=True)
    late_payment_interest = fields.Float(string="Late Payment Interest %",
                                         tracking=True)
    late_payment_interest_amt = fields.Monetary(
        string='Late Payment Interest Amount', tracking=True,
        compute='_get_interest_amt',
        store=True)
    monthly_interest_rate = fields.Float(string="Monthly Interest Rate %",
                                         tracking=True)
    presently_rate = fields.Float(string="Presently Rate %", tracking=True)
    tag_ids = fields.Many2many('property.tag', string='Tags', tracking=True)
    active = fields.Boolean(default=True, tracking=True)
    rented_area = fields.Float(string='Rented Area', tracking=True)
    square_area = fields.Selection(related="property_id.square_area",
                                   tracking=True, store=True, readonly=False)
    total_rent_month = fields.Monetary(string='Total Rent/Month',
                                       tracking=True)
    service_charge_month = fields.Monetary(string='Service Charge/Month',
                                           tracking=True)
    rent_incerement = fields.Float(string="Rent Increment %", tracking=True)
    rent_incremnet_period = fields.Integer(string="Rent Increment Period",
                                           tracking=True)
    rent_time_period = fields.Selection([('year', 'Year'), ('month', 'Month')],
                                        string="Rent Period", tracking=True)

    rent_increment_date = fields.Date(string='Rent Increment Date')
    last_increment_date = fields.Date(string='Last Increment Date',
                                      tracking=True)

    is_maintenance_company = fields.Boolean(string="Is Maintenance Company",
                                            related="company_id.is_maintenance_company",
                                            tracking=True)
    is_addendum = fields.Boolean(string="Is Addendum", readonly=True,
                                 tracking=True)
    total_months = fields.Float(compute="_get_months", store=True,
                                string="Months", tracking=True)
    total_rent_amount = fields.Monetary(compute="_get_total_rent_amount",
                                        store=True, tracking=True)
    installment_item_id = fields.Many2one("product.product",
                                          string="Installment Item",
                                          tracking=True,
                                          domain=[('type', '=',
                                                   'service')])
    security_deposit_currency_id = fields.Many2one('res.currency',
                                                   string='Security Deposit Currency',
                                                   tracking=True)
    team_members = fields.Char(tracking=True)
    alert_date = fields.Date(store=True, compute="_get_alert_date")

    parent_property_id = fields.Many2one('parent.property', tracking=True)
    amount_month = fields.Monetary(compute='_compute_amount_month',
                                   tracking=True)

    rent = fields.Selection(
        [('per_square_meter', 'Per Square Meter'), ('fixed', 'Fixed')],
        default='per_square_meter',
        string='Rent Type')
    is_increment = fields.Boolean()
    increment_type = fields.Selection(
        [('percentage', 'Percentage'), ('fixed', 'Fixed')],
        default='percentage')

    resident_id = fields.Many2one(
        'res.partner', 'Resident',
        domain="['|', ('parent_id', '=', tenancy_id), ('id', '=', tenancy_id)]"
    )
    contract_includes = fields.Selection([('rent_only', 'Rent Only'), (
        'rent_and_service', 'Rent & Service Charge')],
                                         'Contract Includes',
                                         default='rent_only')
    service_charge_per_month = fields.Monetary('Service Charge/Month ')
    service_charge_smtr = fields.Monetary('Service Charge/Square Meter')

    @api.depends('contract_type_name', 'property_id')
    def _compute_amount_month(self):
        for rec in self:
            amount_month = 0.0
            if rec.contract_type_name == 'lease' and rec.property_id.type == 'commercial':
                amount_month = rec.total_rent_month
            elif rec.contract_type_name == 'lease' and rec.property_id.type in [
                'land', 'industrial', 'residential']:
                amount_month = rec.total_rent
            elif rec.contract_type_name == 'service' and rec.property_id.type in [
                'land', 'industrial']:
                amount_month = rec.total_rent
            elif rec.contract_type_name == 'service' and rec.property_type in [
                'residential', 'commercial']:
                amount_month = rec.service_charge_month
            rec.amount_month = amount_month

    @api.depends("end_date", "payment_term_id")
    def _get_alert_date(self):
        for rec in self:
            if rec.end_date:
                rec.alert_date = rec.end_date - timedelta(days=45)

    @api.onchange("agreement_template_id")
    def _onchange_agreement_template_get_body(self):
        for rec in self:
            agreement_data = ''
            if rec.agreement_template_id and rec.agreement_template_id.template_variable_ids:
                body = rec.agreement_template_id.agreement
                body_var = set(re.findall(r'{{[0-9][0-9]*}}', body or ''))
                variable_dict = {}
                for var in rec.agreement_template_id.template_variable_ids:
                    variable_dict[var.name] = var.demo_value
                    if var.field_type == 'free_text':
                        variable_dict[
                            var.name] = var.free_text_value if var.free_text_value else var.name
                    elif var.field_type == 'field':
                        variable_dict[var.name] = self.mapped(var.field_name)[
                            0] if self.mapped(
                            var.field_name) else var.name
                for data in body_var:
                    body = body.replace(data, str(variable_dict.get(data)))
                agreement_data = body
            elif rec.agreement_template_id and not rec.agreement_template_id.templare_variable_ids:
                agreement_data = rec.agreement_template_id.agreement
            rec.contract_terms = agreement_data

    @api.onchange("property_id")
    def _onchange_property_id(self):
        self.rent_smtr = self.property_id.rent_smtr
        self.total_rent = self.property_id.tenancy_price
        self.total_rent_month = self.property_id.tenancy_price
        self.rented_area = self.property_id.rented_area
        self.currency_id = self.property_id.currency_id.id
        self.parent_property_id = self.property_id.parent_property_id.id
        self.rent = self.property_id.rent
        if self.contract_type_name == "service" and self.property_type in [
            'residential', 'commercial']:
            self.service_charge_month = self.property_id.tenancy_price

    @api.onchange('property_id', 'rent', 'rented_area', 'service_charge_smtr')
    def onchange_new_service_charge_per_month(self):
        for rec in self:
            if rec.property_id and rec.rent and rec.service_charge_smtr and rec.rent == 'per_square_meter':
                rec.service_charge_per_month = rec.rented_area * rec.service_charge_smtr

    @api.onchange('rent_smtr', 'rented_area', 'contract_type_name')
    def onchange_rent_calculation(self):
        if self.property_type == "commercial" and self.rent == 'per_square_meter' and self.contract_type_name == 'lease':
            self.total_rent_month = self.rent_smtr * self.rented_area
        elif self.property_type in ["commercial",
                                    "residential"] and self.rent == 'per_square_meter' and self.contract_type_name == 'service':
            self.service_charge_month = self.rent_smtr * self.rented_area
        elif self.contract_type_name == "lease" and self.rent == "per_square_meter" and self.property_type in [
            "land",
            "industrial",
            "residential"]:
            self.total_rent = self.rent_smtr * self.rented_area
        elif self.contract_type_name == "service" and self.rent == "per_square_meter" and self.property_type in [
            "land",
            "industrial"]:
            self.total_rent = self.rent_smtr * self.rented_area

    @api.onchange('rent_time_period', 'rent_incremnet_period', 'start_date')
    def onchange_get_rent_increment_date(self):
        unit = self.rent_incremnet_period
        if self.rent_time_period == 'year':
            unit = 12 * self.rent_incremnet_period
        if self.start_date:
            self.rent_increment_date = self.start_date + relativedelta(
                months=unit)

    @api.onchange('property_id', 'contract_type_name')
    def onchange_property_get_company(self):
        if self.contract_type_name == 'service' and self.property_id.maintanance_company_id:
            self.company_id = self.property_id.maintanance_company_id.id
        elif self.contract_type_name == 'lease' and self.property_id.pro_company_id:
            self.company_id = self.property_id.pro_company_id.id

    def unlink(self):
        for rec in self:
            tenancy_rec = self.env["tenancy.details"].sudo().search(
                [('property_id', '=', rec.property_id.id),
                 ('contract_type', '=', 'running_contract'),
                 ('id', '!=', rec.id), ('contract_type_name', '=', 'lease')])
            if rec.contract_type == 'running_contract':
                raise UserError(
                    _('You are only allowed to delete the contract in the Running state !!!'))
            elif rec.contract_type in ['cancel_contract', 'close_contract',
                                       'expire_contract'] and rec.invoice_count > 0:
                raise UserError(
                    _(f"You are not allowed to delete the contract with transactions !!"))
            if not tenancy_rec and rec.property_id.sale_lease == 'for_tenancy':
                rec.property_id.write({"stage": "available"})
            return super(TenancyDetails, self).unlink()

    def toggle_active(self):
        for rec in self:
            if rec.contract_type not in ['cancel_contract', 'close_contract',
                                         'expire_contract']:
                raise UserError(
                    _('You are only allowed to archive the contract in the cancel, close & expire state !!!'))
            return super(TenancyDetails, self).toggle_active()

    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        order = super(TenancyDetails, self).copy(default)
        order.contract_type = 'new_contract'
        return order

    def handover_property(self):
        for rec in self:
            view = self.env.ref(
                'rental_management.handover_property_details_form_view')

            lst = []

            for line in rec.property_id.p_assets_ids:
                lst.append((0, 0, {
                    'serial_no': line.serial_no,
                    'display_type': line.display_type,
                    'product_id': line.product_id,
                    'name': line.name,
                    'Prod_qty': line.Prod_qty,
                    'prod_condition': line.prod_condition,
                    'remark': line.remark,
                    'property_id': line.property_id
                }))

            ctx = {
                'default_contract_id': rec.id,
                'default_property_id': rec.property_id.id,
                'default_main_property_id': rec.property_id.parent_property_id.id,
                'default_tenant_id': rec.tenancy_id.id,
                'default_assets_ids': lst
            }
            return {
                'name': 'Contract',
                'res_model': 'handover.property',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'view_id': view.id,
                'target': 'current',
                'context': ctx
            }

    @api.depends('total_rent', 'late_payment_interest')
    def _get_interest_amt(self):
        for rec in self:
            if rec.total_rent and rec.late_payment_interest:
                rec.late_payment_interest_amt = rec.total_rent * rec.late_payment_interest

    @api.depends("start_date", "end_date")
    def _get_months(self):
        for rec in self:
            if rec.end_date and rec.start_date:
                delta = relativedelta(rec.end_date + timedelta(days=1),
                                      rec.start_date)
                total_months = (delta.years * 12) + delta.months
                total_months = abs(total_months)
                rec.total_months = total_months

    @api.depends("total_rent", "total_months")
    def _get_total_rent_amount(self):
        for rec in self:
            if rec.total_rent and rec.total_months:
                rec.total_rent_amount = rec.total_rent * rec.total_months

    @api.depends('end_date')
    def _get_days(self):
        for rec in self:
            if rec.end_date:
                days = (rec.end_date - fields.Datetime.now().date()).days
                rec.days_left = int(days)

    def create_utility_bill(self):
        for rec in self:
            view = self.env.ref('rental_management.utility_bill_form_view')
            ctx = {
                'default_contract_id': rec.id,
                'default_property_id': rec.property_id.id,
                'default_tenant_name': rec.tenancy_id.id,
            }

            return {
                'name': 'Utility',
                'res_model': 'utility.bill',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'view_id': view.id,
                'target': 'current',
                'context': ctx
            }

    # Sequence Create
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('tenancy_seq', 'New') == 'New':
                vals['tenancy_seq'] = self.env['ir.sequence'].next_by_code(
                    'tenancy.details') or 'New'
        res = super(TenancyDetails, self).create(vals_list)

        return res

    @api.onchange('end_date')
    def onchange_end_date(self):
        for rec in self:
            if rec.start_date:
                if rec.end_date:
                    if rec.end_date < rec.start_date:
                        raise UserError(
                            _("Please select end date bigger than start date !!!"))

    @api.depends('is_any_broker', 'month')
    def _compute_broker_commission(self):
        for rec in self:
            if rec.is_any_broker:
                if rec.rent_type == 'once':
                    if rec.commission_type == 'f':
                        rec.commission = rec.broker_commission
                    else:
                        rec.commission = rec.broker_commission_percentage * rec.total_rent / 100
                else:
                    if rec.commission_type == 'f':
                        rec.commission = rec.broker_commission * rec.month
                    else:
                        rec.commission = rec.broker_commission_percentage * rec.total_rent * rec.month / 100
            else:
                rec.commission = 0

    # Smart Button
    invoice_count = fields.Integer(string='Invoice Count',
                                   compute="_compute_invoice_count")
    handover_count = fields.Integer(string='Handover Count',
                                    compute="_compute_handover_count")
    utility_bill_count = fields.Integer(string='Utility Bill Count',
                                        compute="_compute_utility_count")
    document_count = fields.Integer(
        string='Document Count', compute='_compute_document_count')
    document_ids = fields.Many2many('documents.document', string="Document")

    def _compute_document_count(self):
        for rec in self:
            document_count = self.env['property.documents'].search_count(
                [('tenancy_id', '=', rec.id)])
            rec.document_count = document_count

    def action_property_document_contract(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Document',
            'res_model': 'property.documents',
            'domain': [('tenancy_id', '=', self.id)],
            'context': {'default_tenancy_id': self.id},
            'view_mode': 'list',
            'target': 'current'
        }

    @api.depends('rent_invoice_ids')
    def _compute_invoice_count(self):
        for rec in self:
            count = self.env['rent.invoice'].search_count(
                [('tenancy_id', '=', rec.id)])
            rec.invoice_count = count

    @api.depends('rent_invoice_ids')
    def _compute_handover_count(self):
        for rec in self:
            count = self.env['handover.property'].search_count(
                [('contract_id', '=', rec.id)])
            rec.handover_count = count

    # @api.depends('rent_invoice_ids')
    def _compute_utility_count(self):
        for rec in self:
            count = self.env['utility.bill'].search_count(
                [('contract_id', '=', rec.id)])
            rec.utility_bill_count = count

    def action_handover(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Handover Properties',
            'res_model': 'handover.property',
            'domain': [('contract_id', '=', self.id)],
            'context': {'default_contract_id': self.id},
            'view_mode': 'list,form',
            'target': 'current'
        }

    def action_utility_bill(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Utility Bill',
            'res_model': 'utility.bill',
            'domain': [('contract_id', '=', self.id)],
            'context': {'default_contract_id': self.id},
            'view_mode': 'list,kanban,pivot,graph,form',
            'target': 'current'
        }

    def action_open_contract(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Contract',
            'res_model': 'tenancy.details',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    @api.depends("rent_invoice_ids")
    def _compute_tenancy_calculation(self):
        total = 0.0
        paid = 0.0
        total_tax = 0.0
        for rec in self:
            if rec.rent_invoice_ids:
                for data in rec.rent_invoice_ids:
                    total_tax = total_tax + data.amount_tax
                    total = total + data.rent_invoice_id.amount_total
                    if data.payment_state == "paid":
                        paid = paid + data.rent_invoice_id.amount_total
            rec.total_tenancy = total
            rec.paid_tenancy = paid
            rec.total_tax = total_tax
            rec.remain_tenancy = total - paid

    def action_invoices(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Invoice',
            'res_model': 'rent.invoice',
            'domain': [('tenancy_id', '=', self.id)],
            'view_mode': 'list,form',
            'target': 'current'
        }

    # Button
    def action_close_contract(self):
        tenancy_rec = self.env["tenancy.details"].sudo().search(
            [('property_id', '=', self.property_id.id),
             ('contract_type', '=', 'running_contract'),
             ('id', '!=', self.id), ('contract_type_name', '=', 'lease')])
        self.close_contract_state = True
        if not tenancy_rec and self.property_id.sale_lease == 'for_tenancy':
            self.property_id.write({'stage': 'available'})
        self.contract_type = 'close_contract'
        return True

    # Schedular
    @api.model
    def add_parent_property_to_contract(self):
        contracts = self.env['tenancy.details'].sudo().search([('parent_property_id', '=', False)])
        for rec in contracts:
            if rec.property_id and rec.property_id.parent_property_id:
                rec.parent_property_id = rec.property_id.parent_property_id.id

    @api.model
    def rent_increment_cron(self):
        records = self.env['tenancy.details'].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('rent_incerement', '>', 0.00)])
        today = fields.Date.today()
        for rec in records:
            if rec.rent_increment_date == today:
                if rec.increment_type == 'percentage':
                    if rec.rent == 'per_square_meter':
                        amount = (rec.rent_smtr * rec.rent_incerement) / 100
                        rec.rent_smtr += amount
                        rec.onchange_rent_calculation()
                    else:
                        if rec.property_type in ['land', 'industrial']:
                            amount = (
                                             rec.total_rent * rec.rent_incerement) / 100
                            rec.total_rent += amount
                        elif rec.contract_type_name == 'service' and rec.property_type in [
                            'residential', 'commercial']:
                            amount = (
                                             rec.service_charge_month * rec.rent_incerement) / 100
                            rec.service_charge_month += amount
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'commercial':
                            amount = (
                                             rec.total_rent_month * rec.rent_incerement) / 100
                            rec.total_rent_month += amount
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'residential':
                            amount = (
                                             rec.total_rent * rec.rent_incerement) / 100
                            rec.total_rent += amount
                elif rec.increment_type == 'fixed':
                    if rec.rent == 'per_square_meter':
                        rec.rent_smtr += rec.rent_incerement
                        rec.onchange_rent_calculation()
                    else:
                        if rec.property_type in ['land', 'industrial']:
                            rec.total_rent += rec.rent_incerement
                        elif rec.contract_type_name == 'service' and rec.property_type in [
                            'residential', 'commercial']:
                            rec.service_charge_month += rec.rent_incerement
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'commercial':
                            rec.total_rent_month += rec.rent_incerement
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'residential':
                            rec.total_rent += rec.rent_incerement
                for data in rec.rent_invoice_ids:
                    if data.invoice_date >= rec.rent_increment_date and not data.rent_invoice_id:
                        if rec.property_type in ['land', 'industrial']:
                            data.total_rent = rec.total_rent
                            data.rent_month = data.total_rent
                        elif rec.contract_type_name == 'service' and rec.property_type in [
                            'residential', 'commercial']:
                            data.service_charge_month = rec.service_charge_month
                            data.rent_month = data.service_charge_month
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'commercial':
                            data.rent_per_month = rec.total_rent_month
                            data.rent_month = data.rent_per_month
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'residential':
                            data.total_rent = rec.total_rent
                            data.rent_month = data.total_rent
                        data.onchange_dates_compute_amount()
                unit = rec.rent_incremnet_period
                if rec.rent_time_period == "year":
                    unit = rec.rent_incremnet_period * 12
                rec.rent_increment_date += relativedelta(months=unit)
                rec.last_increment_date = today

    @api.model
    def set_contract_company_cron(self):
        records = self.env['tenancy.details'].sudo().search([])
        for rec in records:
            if rec.contract_type_name == 'lease' and rec.property_id.pro_company_id:
                rec.company_id = rec.property_id.pro_company_id.id
            elif rec.contract_type_name == 'service' and rec.property_id.maintanance_company_id:
                rec.company_id = rec.property_id.maintanance_company_id.id

    @api.model
    def old_contract_invoice_increment(self):
        records = self.env['tenancy.details'].sudo().with_context(skip_check=True).search(
            [('contract_type', '=', 'running_contract'), ('rent_incerement', '>', 0.00)])
        today = fields.Date.today()
        for rec in records:
            if rec.rent_increment_date and rec.rent_increment_date < today:
                if rec.increment_type == 'percentage':
                    if rec.rent == 'per_square_meter':
                        amount = (rec.rent_smtr * rec.rent_incerement) / 100
                        rec.rent_smtr += amount
                        rec.onchange_rent_calculation()
                    else:
                        if rec.property_type in ['land', 'industrial']:
                            amount = (
                                             rec.total_rent * rec.rent_incerement) / 100
                            rec.total_rent += amount
                        elif rec.contract_type_name == 'service' and rec.property_type in [
                            'residential', 'commercial']:
                            amount = (
                                             rec.service_charge_month * rec.rent_incerement) / 100
                            rec.service_charge_month += amount
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'commercial':
                            amount = (
                                             rec.total_rent_month * rec.rent_incerement) / 100
                            rec.total_rent_month += amount
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'residential':
                            amount = (
                                             rec.total_rent * rec.rent_incerement) / 100
                            rec.total_rent += amount
                elif rec.increment_type == 'fixed':
                    if rec.rent == 'per_square_meter':
                        rec.rent_smtr += rec.rent_incerement
                        rec.onchange_rent_calculation()
                    else:
                        if rec.property_type in ['land', 'industrial']:
                            rec.total_rent += rec.rent_incerement
                        elif rec.contract_type_name == 'service' and rec.property_type in [
                            'residential', 'commercial']:
                            rec.service_charge_month += rec.rent_incerement
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'commercial':
                            rec.total_rent_month += rec.rent_incerement
                        elif rec.contract_type_name == 'lease' and rec.property_type == 'residential':
                            rec.total_rent += rec.rent_incerement
                if rec.increment_type:
                    for data in rec.rent_invoice_ids:
                        if data.invoice_date >= rec.rent_increment_date and not data.rent_invoice_id:
                            if rec.property_type in ['land', 'industrial']:
                                data.total_rent = rec.total_rent
                                data.rent_month = data.total_rent
                            elif rec.contract_type_name == 'service' and rec.property_type in [
                                'residential',
                                'commercial']:
                                data.service_charge_month = rec.service_charge_month
                                data.rent_month = data.service_charge_month
                            elif rec.contract_type_name == 'lease' and rec.property_type == 'commercial':
                                data.rent_per_month = rec.total_rent_month
                                data.rent_month = data.rent_per_month
                            elif rec.contract_type_name == 'lease' and rec.property_type == 'residential':
                                data.total_rent = rec.total_rent
                                data.rent_month = data.total_rent
                            data.onchange_dates_compute_amount()
                    unit = rec.rent_incremnet_period
                    if rec.rent_time_period == "year":
                        unit = rec.rent_incremnet_period * 12
                    rec.rent_increment_date += relativedelta(months=unit)
                    rec.last_increment_date = today
                    rec.is_increment = True

    def action_active_contract(self):
        invoice_lines = []
        if self.is_any_broker:
            self.action_broker_invoice()
        self.contract_type = 'running_contract'
        self.active_contract_state = True

        if self.property_type == 'commercial' and self.contract_type == 'lease':
            total_rent_month = self.total_rent_month
        if self.property_type in ['residential',
                                  'commercial'] and self.contract_type_name == 'service':
            total_rent_month = self.service_charge_month
        else:
            total_rent_month = self.total_rent

        if self.payment_term_id.rent_unit == 'Month':
            record = {
                'product_id': self.env.ref(
                    'rental_management.property_product_1').id,
                'name': 'First Invoice of ' + self.property_id.name,
                'quantity': self.payment_term_id.month,
                'price_unit': total_rent_month
            }
            invoice_lines.append((0, 0, record))
            if self.is_any_deposit:
                deposit_record = {
                    'product_id': self.env.ref(
                        'rental_management.property_product_1').id,
                    'name': 'Deposit of ' + self.property_id.name,
                    'quantity': self.payment_term_id.month,
                    'price_unit': self.deposit_amount
                }
                invoice_lines.append((0, 0, deposit_record))
            data = {
                'partner_id': self.tenancy_id.id,
                'move_type': 'out_invoice',
                'invoice_date': fields.Date.today(),
                'invoice_line_ids': invoice_lines,
                'currency_id': self.currency_id.id
            }
            invoice_id = self.env['account.move'].sudo().create(data)
            invoice_id.invoice_period_to_date = self.start_date
            invoice_id.invoice_period_from_date = self.end_date

            invoice_id.tenancy_id = self.id
            invoice_id.tenancy_property_id = self.property_id.id
            invoice_id.tenancy_parent_property_id = self.property_id.parent_property_id.id
            invoice_id.action_post()
            self.last_invoice_payment_date = invoice_id.invoice_date
            # self.sudo().action_send_active_contract()
            amount_total = invoice_id.amount_total

            rent_invoice = {
                'tenancy_id': self.id,
                'type': 'rent',
                'invoice_date': self.start_date,
                'description': 'First Rent',
                'rent_invoice_id': invoice_id.id,
                "rent_month": total_rent_month,
                'amount': total_rent_month * self.payment_term_id.month,
                'rent_amount': total_rent_month if self.property_type == 'residential' and self.contract_type_name == 'service' else self.total_rent * self.payment_term_id.month,
                "company_id": self.company_id.id
            }
            if self.is_any_deposit:
                rent_invoice['description'] = 'First Rent + Deposit'
            else:
                rent_invoice['description'] = 'First Rent'
            self.env['rent.invoice'].create(rent_invoice)

        elif self.payment_term_id.rent_unit == 'Quarter':
            record = {
                'product_id': self.env.ref(
                    'rental_management.property_product_1').id,
                'name': 'First Quarter Invoice of ' + self.property_id.name,
                'quantity': self.payment_term_id.month,
                'price_unit': total_rent_month
            }
            invoice_lines.append((0, 0, record))
            if self.is_any_deposit:
                deposit_record = {
                    'product_id': self.env.ref(
                        'rental_management.property_product_1').id,
                    'name': 'Deposit of ' + self.property_id.name,
                    'quantity': self.payment_term_id.month,
                    'price_unit': self.deposit_amount
                }
                invoice_lines.append((0, 0, deposit_record))
            data = {
                'partner_id': self.tenancy_id.id,
                'move_type': 'out_invoice',
                'invoice_date': fields.Date.today(),
                'invoice_line_ids': invoice_lines,
                'currency_id': self.currency_id.id
            }
            invoice_id = self.env['account.move'].sudo().create(data)
            invoice_id.tenancy_id = self.id
            invoice_id.tenancy_property_id = self.property_id.id
            invoice_id.tenancy_parent_property_id = self.property_id.parent_property_id.id
            invoice_id.invoice_period_to_date = self.start_date
            invoice_id.invoice_period_from_date = self.end_date
            invoice_id.action_post()
            self.last_invoice_payment_date = invoice_id.invoice_date
            # self.action_send_active_contract()
            amount_total = invoice_id.amount_total

            rent_invoice = {
                'tenancy_id': self.id,
                'type': 'rent',
                'invoice_date': self.start_date,
                'description': 'First Quarter Rent',
                'rent_invoice_id': invoice_id.id,
                "rent_month": total_rent_month,
                'amount': total_rent_month * self.payment_term_id.month,
                'rent_amount': self.total_rent * self.payment_term_id.month,
                "company_id": self.company_id.id
            }
            if self.is_any_deposit:
                rent_invoice['description'] = 'First Quarter Rent + Deposit'
            else:
                rent_invoice['description'] = 'First Quarter Rent'
            self.env['rent.invoice'].create(rent_invoice)

        elif self.payment_term_id.rent_unit == "Year":
            if self.is_any_deposit:
                deposit_record = {
                    'product_id': self.env.ref(
                        'rental_management.property_product_1').id,
                    'name': 'Deposit of ' + self.property_id.name,
                    'quantity': self.payment_term_id.month * 12,
                    'price_unit': self.deposit_amount
                }
                invoice_lines.append((0, 0, deposit_record))
            if invoice_lines:
                data = {
                    'partner_id': self.tenancy_id.id,
                    'move_type': 'out_invoice',
                    'invoice_date': fields.Date.today(),
                    'invoice_line_ids': invoice_lines,
                    'tenancy_id': self.id,
                    'currency_id': self.currency_id.id,
                    "tenancy_property_id": self.property_id.id,
                    "tenancy_parent_property_id": self.property_id.parent_property_id.id,
                }
                inv_id = self.env['account.move'].sudo().create(data)

                inv_id.invoice_period_to_date = self.start_date
                inv_id.invoice_period_from_date = self.end_date
                inv_id.action_post()

                rent_invoice = {
                    'tenancy_id': self.id,
                    'type': 'rent',
                    'invoice_date': self.start_date,
                    'description': 'First Yearly Rent',
                    'rent_invoice_id': inv_id.id,
                    "rent_month": total_rent_month,
                    'amount': total_rent_month * (
                            self.payment_term_id.month * 12),
                    'rent_amount': inv_id.amount_total * (
                            self.payment_term_id.month * 12),
                    "company_id": self.company_id.id
                }

                if self.is_any_deposit and self.is_extra_service:
                    rent_invoice['description'] = "Deposit + Extra Service"
                elif self.is_any_deposit:
                    rent_invoice['description'] = "Deposit Amount"
                elif self.is_extra_service:
                    rent_invoice['description'] = "Extra Service"
                self.env['rent.invoice'].create(rent_invoice)

            line = []
            record = {
                'product_id': self.env.ref(
                    'rental_management.property_product_1').id,
                'name': 'First Year Invoice of ' + self.property_id.name,
                'quantity': self.payment_term_id.month * 12,
                'price_unit': total_rent_month
            }
            line.append((0, 0, record))
            data = {
                'partner_id': self.tenancy_id.id,
                'move_type': 'out_invoice',
                'invoice_date': fields.Date.today(),
                'invoice_line_ids': line,
                'tenancy_id': self.id,
                'currency_id': self.currency_id.id,
                "tenancy_property_id": self.property_id.id,
                "tenancy_parent_property_id": self.property_id.parent_property_id.id,
            }
            invoice = self.env['account.move'].sudo().create(data)
            invoice.invoice_period_to_date = self.start_date
            invoice.invoice_period_from_date = self.end_date
            invoice.action_post()

            self.last_invoice_payment_date = invoice.invoice_date

            rent_invoice = {
                'tenancy_id': self.id,
                'is_yearly': True,
                'type': 'rent',
                'invoice_date': self.start_date,
                'description': 'First Year Rent',
                'rent_invoice_id': invoice.id,
                "rent_month": total_rent_month,
                'amount': total_rent_month * (self.payment_term_id.month * 12),
                'rent_amount': self.total_rent * 12,
                "company_id": self.company_id.id
            }
            self.env['rent.invoice'].create(rent_invoice)

    def action_cancel_contract(self):
        tenancy_rec = self.env["tenancy.details"].sudo().search(
            [('property_id', '=', self.property_id.id),
             ('contract_type', '=', 'running_contract'),
             ('id', '!=', self.id),
             ('contract_type_name', '=', 'lease')])
        if self.invoice_count > 0:
            raise UserError(
                _(f"You are not allowed to Cancel the Contract with transactions !!"))
        self.close_contract_state = True
        if not tenancy_rec and self.property_id.sale_lease == 'for_tenancy':
            self.property_id.write({'stage': 'available'})
        self.contract_type = 'cancel_contract'

    def action_send_expiry_alert(self):
        """contract is going to expire in 90 days"""
        today = fields.Date.today()
        records = self.env["tenancy.details"].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('contract_type_name', '=', 'lease')])
        mail_template = self.env.ref(
            "rental_management.agreement_expiry_alert_mail_template").sudo()
        sender_email = self.env['ir.config_parameter'].sudo().get_param(
            'rental_management.alert_mail_sender_email')
        for rec in records:
            if rec.end_date:
                remaining_days = (rec.end_date - today).days
                if remaining_days in (90, 60, 30) and mail_template and sender_email:
                    company = rec.company_id
                    mails = company.agreement_expiring_in_ninty_days_ids.mapped(
                        'email')
                    filtered_mails = [item for item in mails if
                                      not isinstance(item, bool)]
                    mails_str = ""
                    if filtered_mails:
                        mails_str = ", ".join(filtered_mails)
                    mail_values = {
                        "email_cc": mails_str,
                        "email_from": sender_email,
                        "subject": f"Notice: Lease Agreement Expiry in {remaining_days} Days [{rec.property_id.name} - {rec.property_id.parent_property_id.name}]"
                    }
                    ctx = {
                        'remaining_days': remaining_days
                    }
                    mail_template.with_context(ctx).send_mail(rec.id, force_send=True,
                                                              email_values=mail_values)

    def action_send_active_contract(self):
        mail_template = self.env.ref(
            'rental_management.active_contract_mail_template').sudo()
        if mail_template:
            mail_template.send_mail(self.id, force_send=True)

    def action_send_tenancy_reminder(self):
        mail_template = self.env.ref(
            'rental_management.tenancy_reminder_mail_template').sudo()
        if mail_template:
            mail_template.send_mail(self.id, force_send=True)

    def action_expiry_alert_list_xls(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Expiring Contracts in 45 Days"
        today = fields.Date.today()
        records = self.env["tenancy.details"].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('contract_type_name', '=', 'lease')])

        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        headers = ["Ref.", "Contract Type", "Property", "Tenant", "Landlord", "Start Date", "End Date"]
        col_widths = [15, 25, 30, 20, 20, 15, 15]
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        row = 2
        for rec in records:
            if (rec.end_date - today).days <= 45:
                contract_name = "Lease Agreement" if rec.contract_type_name == "lease" else "Property Management Agreement"
                values = [
                    rec.tenancy_seq,
                    contract_name,
                    f"{rec.property_id.name} - {rec.property_id.type}",
                    rec.tenancy_id.name,
                    rec.property_landlord_id.name,
                    rec.start_date.strftime('%m/%d/%Y') if rec.start_date else '',
                    rec.end_date.strftime('%m/%d/%Y') if rec.end_date else '',
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = sheet.cell(row=row, column=col_idx, value=val)
                    cell.alignment = center_align
                row += 1

        stream = BytesIO()
        workbook.save(stream)
        out = base64.encodebytes(stream.getvalue())

        attachment = self.env['ir.attachment'].sudo()
        today = fields.Date.today().strftime("%m/%d/%Y")

        filename = today + " Expiring Contracts.xlsx"
        attachment_id = attachment.create(
            {'name': filename,
             'type': 'binary',
             'public': False,
             'datas': out})
        return attachment_id

    def send_expiring_contract_list(self):
        """contract expiring list to team members"""
        today = fields.Date.today()
        records = self.env["tenancy.details"].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('alert_date', '<=', today),
             ('contract_type_name', '=', 'lease')], limit=1)
        mail_template = self.env.ref(
            "rental_management.contract_expiry_list_alert_mail_template").sudo()
        sender_email = self.env['ir.config_parameter'].sudo().get_param(
            'rental_management.alert_mail_sender_email')
        attachment_id = self.action_expiry_alert_list_xls()
        company = records.company_id
        mail_ids = company.agreement_expiring_in_forty_five_days.mapped('id')
        names = company.agreement_expiring_in_forty_five_days.mapped('name')
        filtered_names = [item for item in names if
                          not isinstance(item, bool)]
        names_str = ""
        if filtered_names:
            names_str = ", ".join(filtered_names)
        mail_values = {
            'attachment_ids': [(6, 0, [attachment_id.id])],
            'recipient_ids': [(6, 0, mail_ids)],
            # "email_to": mail_str,
            'email_from': sender_email
        }
        if records and mail_template and sender_email:
            mail_template.with_context({"team_member": names_str}).send_mail(
                records.id, email_values=mail_values,
                force_send=True)

    def action_broker_invoice(self):
        record = {
            'product_id': self.env.ref(
                'rental_management.property_product_1').id,
            'name': 'Brokerage of ' + self.property_id.name,
            'quantity': 1,
            'price_unit': self.commission
        }
        invoice_lines = [(0, 0, record)]
        data = {
            'partner_id': self.tenancy_id.id,
            'move_type': 'in_invoice',
            'invoice_date': fields.Date.today(),
            'invoice_line_ids': invoice_lines
        }
        invoice_id = self.env['account.move'].sudo().create(data)
        invoice_id.tenancy_id = self.id
        invoice_id.tenancy_property_id = self.property_id.id
        invoice_id.tenancy_parent_property_id = self.property_id.parent_property_id.id
        invoice_id.action_post()
        self.broker_invoice_state = True
        self.broker_invoice_id = invoice_id.id
        return True

    @api.model
    def tenancy_recurring_invoice(self):
        today_date = fields.Date.today()
        reminder_days = self.env['ir.config_parameter'].sudo().get_param(
            'rental_management.reminder_days')
        tenancy_contracts = self.env['tenancy.details'].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('payment_term_id.duration', '=', 'Monthly'),
             ('rent_unit', '=', 'Month'), ('contract_type_name', '=', 'lease'),
             ('type', '=', 'automatic')])
        for rec in tenancy_contracts:
            if rec.contract_type == 'running_contract' and rec.payment_term_id.duration == 'Monthly':
                if today_date < rec.end_date:
                    invoice_date = rec.last_invoice_payment_date + relativedelta(
                        months=1)
                    next_invoice_date = rec.last_invoice_payment_date + relativedelta(
                        months=1) - relativedelta(
                        days=int(reminder_days))
                    if today_date == next_invoice_date:
                        record = {
                            'product_id': self.env.ref(
                                'rental_management.property_product_1').id,
                            'name': 'Installment of ' + rec.property_id.name,
                            'quantity': 1,
                            'price_unit': rec.total_rent
                        }
                        invoice_lines = [(0, 0, record)]
                        if rec.is_extra_service:
                            for line in rec.extra_services_ids:
                                if line.service_type == "monthly":
                                    desc = "Monthly"
                                    service_invoice_record = {
                                        'product_id': line.service_id.id,
                                        'name': desc,
                                        'quantity': 1,
                                        'price_unit': line.price
                                    }
                                    invoice_lines.append(
                                        (0, 0, service_invoice_record))
                        data = {
                            'partner_id': rec.tenancy_id.id,
                            'move_type': 'out_invoice',
                            'invoice_date': invoice_date,
                            'invoice_line_ids': invoice_lines
                        }
                        invoice_id = self.env['account.move'].sudo().create(
                            data)
                        invoice_id.tenancy_id = rec.id
                        invoice_id.tenancy_property_id = rec.property_id.id
                        invoice_id.tenancy_parent_property_id = rec.property_id.parent_property_id.id
                        invoice_id.action_post()
                        rec.last_invoice_payment_date = invoice_id.invoice_date
                        rent_invoice = {
                            'tenancy_id': rec.id,
                            'type': 'rent',
                            'invoice_date': invoice_date,
                            'description': 'Installment of ' + rec.property_id.name,
                            'rent_invoice_id': invoice_id.id,
                            'amount': invoice_id.amount_total,
                            'rent_amount': rec.total_rent,
                            "company_id": rec.company_id.id
                        }
                        self.env['rent.invoice'].create(rent_invoice)
                        rec.action_send_tenancy_reminder()

    @api.model
    def tenancy_expire(self):
        today_date = fields.Date.today()
        tenancy_contracts = self.env['tenancy.details'].sudo().search(
            [('contract_type', 'in', ['running_contract', 'expire_contract'])])
        for rec in tenancy_contracts:
            tenancy_rec = self.env["tenancy.details"].sudo().search(
                [('property_id', '=', rec.property_id.id),
                 ('contract_type', '=', 'running_contract'),
                 ('contract_type_name', '=', 'lease'),
                 ('id', '!=', rec.id)], limit=1)
            if today_date > rec.end_date:
                rec.contract_type = 'expire_contract'
                if not tenancy_rec and rec.property_id.sale_lease == 'for_tenancy':
                    rec.property_id.stage = 'available'

    @api.model
    def property_stage_update(self):
        tenancy_contracts = self.env['tenancy.details'].sudo().search(
            [('contract_type', '=', 'expire_contract')])
        for rec in tenancy_contracts:
            tenancy_rec = self.env["tenancy.details"].sudo().search(
                [('property_id', '=', rec.property_id.id),
                 ('contract_type', '=', 'running_contract'),
                 ('contract_type_name', '=', 'lease')])
            if not tenancy_rec and rec.property_id.sale_lease == 'for_tenancy':
                rec.property_id.stage = "available"

    @api.model
    def tenancy_recurring_quarterly_invoice(self):
        today_date = fields.Date.today()
        # today_date = datetime.date(2024, 4, 1)
        reminder_days = self.env['ir.config_parameter'].sudo().get_param(
            'rental_management.reminder_days')
        tenancy_contracts = self.env['tenancy.details'].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('payment_term_id.duration', '=', 'Quarterly'),
             ('type', '=', 'automatic'), ('rent_unit', '=', 'Month'),
             ('contract_type_name', '=', 'lease')])
        for rec in tenancy_contracts:
            if rec.contract_type == 'running_contract' and rec.payment_term_id.duration == 'Quarterly':
                if today_date < rec.end_date:
                    invoice_date = rec.last_invoice_payment_date + relativedelta(
                        months=3)
                    next_next_invoice_date = invoice_date + relativedelta(
                        months=3)
                    next_invoice_date = rec.last_invoice_payment_date + relativedelta(
                        months=3) - relativedelta(
                        days=int(reminder_days))
                    if rec.end_date < next_next_invoice_date:
                        delta = relativedelta(next_next_invoice_date,
                                              rec.end_date)
                        diff = delta.months
                    else:
                        diff = 0
                    if today_date == next_invoice_date:
                        record = {
                            'product_id': self.env.ref(
                                'rental_management.property_product_1').id,
                            'name': 'Quarterly Installment of ' + rec.property_id.name,
                            'quantity': 1,
                            'price_unit': rec.total_rent * (3 - diff)
                        }
                        invoice_lines = [(0, 0, record)]
                        if rec.is_extra_service:
                            for line in rec.extra_services_ids:
                                if line.service_type == "monthly":
                                    desc = "Quarterly Service"
                                    service_invoice_record = {
                                        'product_id': line.service_id.id,
                                        'name': desc,
                                        'quantity': 1,
                                        'price_unit': line.price * (3 - diff)
                                    }
                                    invoice_lines.append(
                                        (0, 0, service_invoice_record))
                        data = {
                            'partner_id': rec.tenancy_id.id,
                            'move_type': 'out_invoice',
                            'invoice_date': invoice_date,
                            'invoice_line_ids': invoice_lines
                        }
                        invoice_id = self.env['account.move'].sudo().create(
                            data)
                        invoice_id.tenancy_id = rec.id
                        invoice_id.tenancy_property_id = rec.property_id.id
                        invoice_id.tenancy_parent_property_id = rec.property_id.parent_property_id.id
                        invoice_id.action_post()
                        rec.last_invoice_payment_date = invoice_id.invoice_date
                        rent_invoice = {
                            'tenancy_id': rec.id,
                            'type': 'rent',
                            'invoice_date': invoice_date,
                            'description': 'Quarterly Installment of ' + rec.property_id.name,
                            'rent_invoice_id': invoice_id.id,
                            'amount': invoice_id.amount_total,
                            'rent_amount': rec.total_rent * (3 - diff),
                            "company_id": rec.company_id.id
                        }
                        self.env['rent.invoice'].create(rent_invoice)
                        rec.action_send_tenancy_reminder()

    @api.model
    def tenancy_yearly_invoice(self):
        today_date = fields.Date.today()
        reminder_days = self.env['ir.config_parameter'].sudo().get_param(
            'rental_management.reminder_days')
        tenancy_contracts = self.env['tenancy.details'].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('type', '=', 'automatic'),
             ('payment_term_id.duration', '=', 'Year'),
             ('rent_unit', '=', 'Year'), ('contract_type_name', '=', 'lease')])
        for rec in tenancy_contracts:
            if today_date < rec.end_date:
                invoice_date = rec.last_invoice_payment_date + relativedelta(
                    years=1)
                next_invoice_date = rec.last_invoice_payment_date + relativedelta(
                    years=1) - relativedelta(
                    days=int(reminder_days))
                if today_date == next_invoice_date:
                    record = {
                        'product_id': self.env.ref(
                            'rental_management.property_product_1').id,
                        'name': 'Yearly Installment of ' + rec.property_id.name,
                        'quantity': 1,
                        'price_unit': rec.total_rent
                    }
                    invoice_lines = [(0, 0, record)]
                    if rec.is_extra_service:
                        for line in rec.extra_services_ids:
                            if line.service_type == "monthly":
                                desc = "Monthly"
                                service_invoice_record = {
                                    'product_id': line.service_id.id,
                                    'name': desc,
                                    'quantity': 12,
                                    'price_unit': line.price
                                }
                                invoice_lines.append(
                                    (0, 0, service_invoice_record))
                    data = {
                        'partner_id': rec.tenancy_id.id,
                        'move_type': 'out_invoice',
                        'invoice_date': invoice_date,
                        'invoice_line_ids': invoice_lines
                    }
                    invoice_id = self.env['account.move'].sudo().create(data)
                    invoice_id.tenancy_id = rec.id
                    invoice_id.tenancy_property_id = rec.property_id.id
                    invoice_id.tenancy_parent_property_id = rec.property_id.parent_property_id.id
                    invoice_id.action_post()
                    rec.last_invoice_payment_date = invoice_id.invoice_date
                    rent_invoice = {
                        'tenancy_id': rec.id,
                        'type': 'rent',
                        'invoice_date': invoice_date,
                        'description': 'Installment of ' + rec.property_id.name,
                        'rent_invoice_id': invoice_id.id,
                        'amount': invoice_id.amount_total,
                        'rent_amount': rec.total_rent,
                        "company_id": rec.company_id.id
                    }
                    self.env['rent.invoice'].create(rent_invoice)
                    rec.action_send_tenancy_reminder()

    @api.model
    def tenancy_manual_invoice(self):
        today_date = fields.Date.today()
        reminder_days = self.env['ir.config_parameter'].sudo().get_param(
            'rental_management.reminder_days')
        tenancy_contracts = self.env['tenancy.details'].sudo().search(
            [('contract_type', '=', 'running_contract'),
             ('type', '=', 'manual'), ('contract_type_name', '=', 'lease')])
        for data in tenancy_contracts:
            for rec in data.rent_invoice_ids:
                if not rec.rent_invoice_id:
                    invoice_date = rec.invoice_date - relativedelta(
                        days=int(reminder_days))
                    if today_date == invoice_date:
                        rec.action_create_invoice()
            data.action_send_tenancy_reminder()


class ContractDuration(models.Model):
    _name = 'contract.duration'
    _description = 'Contract Duration and Month'
    _rec_name = 'duration'

    duration = fields.Char(string='Duration', required=True)
    month = fields.Integer(string='Unit')
    rent_unit = fields.Selection([('Month', "Month"),
                                  ('Quarter', "Quarter"),
                                  ('Year', "Year")],
                                 default='Month',
                                 string="Rent Unit")


class TenancyExtraServiceLine(models.Model):
    _name = "tenancy.service.line"
    _description = "Tenancy Service Line"

    service_id = fields.Many2one('product.product', string="Service", domain=[
        ('is_extra_service_product', '=', True)])
    price = fields.Float(related="service_id.lst_price", string="Cost")
    service_type = fields.Selection([('once', 'Once'), ('monthly', 'Monthly')],
                                    string="Type", default="once")
    tenancy_id = fields.Many2one('tenancy.details', string="Tenancies")
    from_contract = fields.Boolean()

    def action_create_service_invoice(self):
        self.from_contract = True
        record = {
            'product_id': self.service_id.id,
            'name': "Extra Added Service",
            'quantity': 1,
            'price_unit': self.service_id.lst_price
        }
        invoice_lines = [(0, 0, record)]
        data = {
            'partner_id': self.tenancy_id.tenancy_id.id,
            'move_type': 'out_invoice',
            'invoice_date': fields.Date.today(),
            'invoice_line_ids': invoice_lines
        }
        invoice_id = self.env['account.move'].sudo().create(data)
        invoice_id.tenancy_id = self.tenancy_id.id
        invoice_id.tenancy_property_id = self.tenancy_id.property_id.id
        invoice_id.tenancy_parent_property_id = self.tenancy_id.property_id.parent_property_id.id
        invoice_id.action_post()
        rent_invoice = {
            'tenancy_id': self.tenancy_id.id,
            'type': 'maintenance',
            'amount': self.service_id.lst_price,
            'invoice_date': fields.Date.today(),
            'description': 'New Service',
            'rent_invoice_id': invoice_id.id,
            "company_id": self.tenancy_id.company_id.id
        }
        self.env['rent.invoice'].create(rent_invoice)


class AgreementTemplate(models.Model):
    _name = "agreement.template"
    _description = "Agreement Template"

    name = fields.Char(string="Title")
    agreement = fields.Html(string="Agreement")
    model = fields.Char(default="tenancy.details")
    template_variable_ids = fields.One2many('agreement.template.variables',
                                            'template_id',
                                            string="Template Variables",
                                            store=True,
                                            compute='_compute_variable_ids',
                                            precompute=True,
                                            readonly=False)

    @api.depends('agreement')
    def _compute_variable_ids(self):
        for rec in self:
            delete_var = self.env['agreement.template.variables']
            keep_var = self.env['agreement.template.variables']
            created_var = []
            body_var = set(re.findall(r'{{[0-9][0-9]*}}', rec.agreement or ''))
            existing_var = rec.template_variable_ids
            new_var = [var_name for var_name in body_var if
                       var_name not in existing_var.mapped('name')]
            deleted_var = existing_var.filtered(
                lambda var: var.name not in body_var)
            created_var += [{'name': var_name} for var_name in set(new_var)]
            delete_var += deleted_var
            keep_var += existing_var - deleted_var
            rec.template_variable_ids = [(3, to_remove.id) for to_remove in
                                         delete_var] + [(0, 0, vals) for vals
                                                        in
                                                        created_var]


class AgreementTemplateVariables(models.Model):
    """Variables for dynamic agreement template body"""
    _name = "agreement.template.variables"
    _description = __doc__

    name = fields.Char(string='Name')
    demo_value = fields.Char(string='Sample Value', default='Sample Value')
    field_type = fields.Selection(
        [('free_text', 'Free Text'), ('field', 'Field of Model')],
        string='Type')
    field_name = fields.Char(string='Field')
    template_id = fields.Many2one('agreement.template')
    model = fields.Char(related='template_id.model')
    free_text_value = fields.Char(string='Free Text Value')
