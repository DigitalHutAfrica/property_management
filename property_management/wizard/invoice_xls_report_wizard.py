# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from odoo import models


class InvoiceExcelReport(models.TransientModel):
    _name = "invoice.excel.report"
    _description = "Print all invoices with details in excel file"

    def print_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rent Invoice"

        bold = Font(bold=True)
        center = Alignment(horizontal='center')
        right = Alignment(horizontal='right')

        headings = [
            "Invoice No", "Invoice Date", "Property", "Tenant",
            "Contract Start Date", "Contract End Date", "Contract Months",
            "Invoice From", "Invoice To", "Invoice Months",
            "Rented Area(Sq.M)", "Rent/Square Meter", "Rent/Month", "Total Rent",
            "VAT @ 18%", "Total", "Currency Rate", "Amount in TSH",
            "VAT TSH", "Total Amount TSH"
        ]
        for col, heading in enumerate(headings, 1):
            cell = ws.cell(row=1, column=col, value=heading)
            cell.font = bold
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = 18

        domain = [("rent_invoice_id", "!=", False), ("tenancy_id.contract_type_name", "=", "lease")]
        records = self.env["rent.invoice"].search(domain)

        row = 2
        for rec in records:
            invoice_months = self.get_invoice_months(rec.invoice_period_from_date, rec.invoice_period_to_date) if rec.invoice_period_to_date and rec.invoice_period_from_date else ""
            contract_months = self.get_invoice_months(rec.tenancy_id.end_date, rec.tenancy_id.start_date) if rec.tenancy_id.start_date and rec.tenancy_id.end_date else ""
            rate = 0
            if rec.rent_invoice_id.amount_total:
                rate = rec.rent_invoice_id.amount_total_signed / rec.rent_invoice_id.amount_total

            values = [
                rec.rent_invoice_id.name,
                rec.invoice_date.strftime('%d/%m/%Y') if rec.invoice_date else '',
                rec.property_id.name,
                rec.customer_id.name,
                rec.tenancy_id.start_date.strftime('%d/%m/%Y') if rec.tenancy_id.start_date else '',
                rec.tenancy_id.end_date.strftime('%d/%m/%Y') if rec.tenancy_id.end_date else '',
                contract_months,
                rec.invoice_period_to_date.strftime('%d/%m/%Y') if rec.invoice_period_to_date else '',
                rec.invoice_period_from_date.strftime('%d/%m/%Y') if rec.invoice_period_from_date else '',
                invoice_months,
                rec.tenancy_id.rented_area,
                rec.tenancy_id.rent_smtr,
                f"{rec.currency_id.symbol} {rec.rent}",
                f"{rec.currency_id.symbol} {rec.amount}",
                f"{rec.currency_id.symbol} {rec.rent_invoice_id.amount_tax}",
                f"{rec.currency_id.symbol} {rec.rent_invoice_id.amount_total}",
                str(rate),
                rec.rent_invoice_id.amount_untaxed_signed,
                rec.rent_invoice_id.amount_tax_signed,
                rec.rent_invoice_id.amount_total_signed,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = right if col > 2 else center
            row += 1

        stream = BytesIO()
        wb.save(stream)
        out = base64.encodebytes(stream.getvalue())
        attachment_id = self.env['ir.attachment'].sudo().create({
            'name': "Invoice Master (Rental Property).xlsx",
            'type': 'binary', 'public': False, 'datas': out
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment_id.id,
            'target': 'self',
        }

    def get_invoice_months(self, end_date, start_date):
        if not end_date or not start_date:
            return 0
        delta = relativedelta(end_date + timedelta(days=1), start_date)
        return abs((delta.years * 12) + delta.months)
