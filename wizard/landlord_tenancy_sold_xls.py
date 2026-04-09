# -*- coding: utf-8 -*-
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from odoo import fields, models


# Color fills for payment status
FILL_RED     = PatternFill("solid", fgColor="FF4444")
FILL_GREEN   = PatternFill("solid", fgColor="44BB44")
FILL_MAGENTA = PatternFill("solid", fgColor="CC44CC")
FILL_GOLD    = PatternFill("solid", fgColor="FFD700")
FILL_VIOLET  = PatternFill("solid", fgColor="8844AA")
FILL_BLUE    = PatternFill("solid", fgColor="6688AA")
FILL_HEADER  = PatternFill("solid", fgColor="DDDDDD")

BOLD_CENTER = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT  = Alignment(horizontal='right', vertical='center')


def _status_fill(payment_state):
    return {
        'paid': FILL_GREEN, 'not_paid': FILL_RED,
        'reversed': FILL_MAGENTA, 'partial': FILL_BLUE,
        'in_payment': FILL_VIOLET,
    }.get(payment_state, FILL_GOLD)


def _status_label(payment_state):
    return {
        'paid': 'Paid', 'not_paid': 'Not Paid', 'reversed': 'Reversed',
        'partial': 'Partial Paid', 'in_payment': 'In Payment',
    }.get(payment_state, 'Invoicing App Legacy')


def _stage_label(contract_type):
    return {
        'new_contract': 'Draft', 'running_contract': 'Running',
        'cancel_contract': 'Cancel', 'close_contract': 'Close',
    }.get(contract_type, 'Expire')


def _write_tenancy_headers(ws, cols):
    for col, header in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD_CENTER
        cell.alignment = CENTER
        cell.fill = FILL_HEADER
        ws.column_dimensions[get_column_letter(col)].width = 20


def _write_tenancy_row(ws, row, data, payment_state):
    fill = _status_fill(payment_state)
    for col, (val, align) in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = align
    # colour the payment status column (second to last)
    ws.cell(row=row, column=len(data) - 1).fill = fill


class LandlordSaleTenancy(models.TransientModel):
    _name = 'landlord.sale.tenancy'
    _description = "Landlord Tenancy And sale Report"
    _rec_name = "landlord_id"

    landlord_id = fields.Many2one('res.partner', domain="[('is_landlord','=',True)]")
    report_for = fields.Selection([('tenancy', 'Tenancy'), ('sold', 'Property Sold')], string="Report For")

    def _save(self, wb, filename):
        stream = BytesIO()
        wb.save(stream)
        out = base64.encodebytes(stream.getvalue())
        att = self.env['ir.attachment'].sudo().create({
            'name': filename, 'type': 'binary', 'public': False, 'datas': out
        })
        return {'type': 'ir.actions.act_url', 'url': '/web/content/%s?download=true' % att.id, 'target': 'self'}

    def action_tenancy_sold_xls_report(self):
        if self.report_for == "tenancy":
            return self._tenancy_report()
        elif self.report_for == "sold":
            return self._sold_report()

    def _tenancy_report(self):
        name = "Tenancy Information - " + self.landlord_id.name
        wb = openpyxl.Workbook()
        sheets = {
            'all': wb.active,
            'paid': wb.create_sheet('Paid Tenancies'),
            'not_paid': wb.create_sheet('Not Paid Tenancy'),
            'partial': wb.create_sheet('Partial Paid Tenancies'),
        }
        sheets['all'].title = 'Landlord wise Tenancies'

        all_headers = ['Date', 'Tenancy No.', 'Tenant', 'Property', 'Invoice Ref.', 'Payment Term', 'Amount', 'Payment Status', 'Tenancy Status']
        sub_headers = ['Date', 'Tenancy No.', 'Tenant', 'Property', 'Invoice Ref.', 'Amount', 'Payment Status', 'Tenancy Status']

        for ws in sheets.values():
            hdrs = all_headers if ws == sheets['all'] else sub_headers
            _write_tenancy_headers(ws, hdrs)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdrs))
            ws.cell(row=1, column=1).value = name
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = CENTER

        # re-write sub-headers at row 2
        for ws in sheets.values():
            hdrs = all_headers if ws == sheets['all'] else sub_headers
            for col, h in enumerate(hdrs, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.font = BOLD_CENTER
                cell.alignment = CENTER
                cell.fill = FILL_HEADER

        rows = {k: 3 for k in sheets}
        rent_invoices = self.env['rent.invoice'].search([('landlord_id', '=', self.landlord_id.id)])
        for data in rent_invoices:
            ps = data.payment_state or ''
            amount = f"{data.rent_invoice_id.amount_total} {data.currency_id.symbol}"
            pt = data.tenancy_id.payment_term_id.duration or ''
            stage = _stage_label(data.tenancy_id.contract_type)
            status = _status_label(ps)
            fill = _status_fill(ps)
            date_str = data.invoice_date.strftime('%d/%m/%Y') if data.invoice_date else ''

            all_row = [date_str, data.tenancy_id.tenancy_seq, data.tenancy_id.tenancy_id.name,
                       data.tenancy_id.property_id.name, data.rent_invoice_id.name, pt, amount, status, stage]
            sub_row = [date_str, data.tenancy_id.tenancy_seq, data.tenancy_id.tenancy_id.name,
                       data.tenancy_id.property_id.name, data.rent_invoice_id.name, amount, status, stage]

            def write_row(ws, row_data, r):
                for col, val in enumerate(row_data, 1):
                    ws.cell(row=r, column=col, value=val).alignment = CENTER
                status_col = 8 if len(row_data) == 9 else 7
                ws.cell(row=r, column=status_col).fill = fill

            write_row(sheets['all'], all_row, rows['all'])
            rows['all'] += 1

            if ps == 'paid':
                write_row(sheets['paid'], sub_row, rows['paid'])
                rows['paid'] += 1
            elif ps == 'not_paid':
                write_row(sheets['not_paid'], sub_row, rows['not_paid'])
                rows['not_paid'] += 1
            elif ps == 'partial':
                write_row(sheets['partial'], sub_row, rows['partial'])
                rows['partial'] += 1

        return self._save(wb, self.landlord_id.name + ".xlsx")

    def _sold_report(self):
        name = "Sold Information - " + self.landlord_id.name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Landlord wise Sold Information'
        headers = ['Date', 'Sequence', 'Customer', 'Property', 'Sale Price', 'Invoice Reference', 'Payment Status', 'Sold Status']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).value = name
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = CENTER
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = BOLD_CENTER
            cell.alignment = CENTER
            cell.fill = FILL_HEADER
            ws.column_dimensions[get_column_letter(col)].width = 20
        row = 3
        for data in self.env['property.vendor'].search([('landlord_id', '=', self.landlord_id.id)]):
            ps = data.sold_invoice_payment_state or ''
            amount = f"{data.sale_price} {data.currency_id.symbol}"
            stage = {'booked': 'Booked', 'refund': 'Refund', 'sold': 'Sold'}.get(data.stage, '')
            status = _status_label(ps)
            fill = _status_fill(ps)
            date_str = data.date.strftime('%d/%m/%Y') if data.date else ''
            vals = [date_str, data.sold_seq, data.customer_id.name, data.property_id.name,
                    amount, data.sold_invoice_id.name or '', status, stage]
            for col, val in enumerate(vals, 1):
                ws.cell(row=row, column=col, value=val).alignment = CENTER
            ws.cell(row=row, column=7).fill = fill
            row += 1
        return self._save(wb, self.landlord_id.name + ".xlsx")
