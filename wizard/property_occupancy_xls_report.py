# -*- coding: utf-8 -*-
import base64
import calendar
from io import BytesIO
from datetime import timedelta, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from odoo import fields, models


RED_FILL    = PatternFill("solid", fgColor="FF6666")
BOLD_CENTER = Font(bold=True)
CENTER      = Alignment(horizontal='center', vertical='center')
HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")


def get_dates_between(start_date, end_date):
    result, current = set(), start_date
    while current <= end_date:
        result.add(current)
        current += timedelta(days=1)
    return result


def get_dates_for_month(year, month):
    num_days = calendar.monthrange(year, month)[1]
    return set(date(year, month, day) for day in range(1, num_days + 1))


class PropertyOccupancyReport(models.TransientModel):
    _name = "property.occupancy.report"
    _description = "Wizard for Property Occupancy Report"

    start_date = fields.Date(string="Start Date")
    end_date   = fields.Date(string="End Date")

    def print_occupancy_report(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Occupancy"

        # Build month list
        month_list = []
        current = self.start_date
        while current <= self.end_date:
            month_list.append(current)
            # advance ~1 month
            current = (current.replace(day=28) + timedelta(days=4)).replace(day=1)

        # Headers
        headers = ["Parent Property", "Property"] + [m.strftime("%m/%y") for m in month_list] + ["Occupied Months", "Vacant Months", "Total Months"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = BOLD_CENTER
            cell.alignment = CENTER
            cell.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 22

        properties = self.env["rent.invoice"].sudo().search(
            [('property_id', '!=', False)]).mapped('property_id')
        invoice_model = self.env['rent.invoice']

        data_row = 2
        for rec in properties:
            parent = rec.parent_property_id.name if rec.parent_property_id else ""
            ws.cell(row=data_row, column=1, value=parent).alignment = CENTER
            ws.cell(row=data_row, column=2, value=rec.name).alignment = CENTER

            occupied_months = 0
            invoices = invoice_model.sudo().search(
                [('property_id', '=', rec.id), ('rent_invoice_id', '!=', False)])

            for m_idx, month_date in enumerate(month_list):
                col = m_idx + 3
                month_dates = get_dates_for_month(month_date.year, month_date.month)
                cell_value = "Vacant"
                for inv in invoices:
                    if inv.tenancy_id.contract_type_name == 'lease' and inv.invoice_period_to_date and inv.invoice_period_from_date:
                        inv_dates = get_dates_between(inv.invoice_period_to_date, inv.invoice_period_from_date)
                        if inv_dates.intersection(month_dates):
                            rent = inv.total_rent if rec.type != 'commercial' else inv.rent_per_month
                            cell_value = f"{inv.currency_id.symbol} {rent}"
                            occupied_months += 1
                            break
                cell = ws.cell(row=data_row, column=col, value=cell_value)
                cell.alignment = CENTER
                if cell_value == "Vacant":
                    cell.fill = RED_FILL

            vacant_months = len(month_list) - occupied_months
            ws.cell(row=data_row, column=len(month_list) + 3, value=occupied_months).alignment = CENTER
            ws.cell(row=data_row, column=len(month_list) + 4, value=vacant_months).alignment = CENTER
            ws.cell(row=data_row, column=len(month_list) + 5, value=len(month_list)).alignment = CENTER
            data_row += 1

        stream = BytesIO()
        wb.save(stream)
        out = base64.encodebytes(stream.getvalue())
        today = fields.Date.today().strftime("%d-%m-%Y")
        att = self.env['ir.attachment'].sudo().create({
            'name': today + " Property Occupancy Report.xlsx",
            'type': 'binary', 'public': False, 'datas': out
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'self',
        }
