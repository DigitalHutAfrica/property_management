# -*- coding: utf-8 -*-
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from odoo import fields, models


class PropertyXlsReport(models.TransientModel):
    _name = 'property.report.wizard'
    _description = 'Create Property Report'
    _rec_name = 'type'

    type = fields.Selection([('tenancy', 'Tenancy'), ('sold', 'Property Sold')], string="Report For")
    start_date = fields.Date(string="Start Date")
    end_date = fields.Date(string="End Date")

    def _make_workbook(self, sheet_name, headers):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        bold = Font(bold=True)
        center = Alignment(horizontal='center')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
            cell.alignment = center
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        return wb, ws

    def _save_attachment(self, wb, filename):
        stream = BytesIO()
        wb.save(stream)
        out = base64.encodebytes(stream.getvalue())
        attachment_id = self.env['ir.attachment'].sudo().create({
            'name': filename, 'type': 'binary', 'public': False, 'datas': out
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment_id.id,
            'target': 'self',
        }

    def action_property_xls_report(self):
        center = Alignment(horizontal='center')
        if self.type == "tenancy":
            headers = ['Tenancy No.', 'Tenant', 'Property', 'Landlord', 'Total Invoiced']
            wb, ws = self._make_workbook('Tenancy Details', headers)
            row = 2
            for group in self.env['account.move'].read_group(
                    [('tenancy_id', '!=', False), ('payment_state', '=', 'paid'),
                     ('invoice_date', '>=', self.start_date), ('invoice_date', '<=', self.end_date)],
                    ['tenancy_id', 'amount_total'], ['tenancy_id'], orderby="amount_total DESC"):
                if group['tenancy_id']:
                    t = self.env['tenancy.details'].sudo().browse(int(group['tenancy_id'][0]))
                    for col, val in enumerate([t.tenancy_seq, t.tenancy_id.name, t.property_id.name,
                                               t.property_landlord_id.name, group['amount_total']], 1):
                        ws.cell(row=row, column=col, value=val).alignment = center
                    row += 1
            return self._save_attachment(wb, 'Tenancy Details.xlsx')

        elif self.type == "sold":
            headers = ['Sequence', 'Customer', 'Property', 'Landlord', 'Total Invoiced']
            wb, ws = self._make_workbook('Property Sold Information', headers)
            row = 2
            for group in self.env['account.move'].read_group(
                    [('sold_id', '!=', False), ('payment_state', '=', 'paid'),
                     ('invoice_date', '>=', self.start_date), ('invoice_date', '<=', self.end_date)],
                    ['sold_id', 'amount_total'], ['sold_id'], orderby="amount_total DESC"):
                if group['sold_id']:
                    s = self.env['property.vendor'].sudo().browse(int(group['sold_id'][0]))
                    for col, val in enumerate([s.sold_seq, s.customer_id.name, s.property_id.name,
                                               s.property_id.landlord_id.name, group['amount_total']], 1):
                        ws.cell(row=row, column=col, value=val).alignment = center
                    row += 1
            return self._save_attachment(wb, 'Sold Information.xlsx')
